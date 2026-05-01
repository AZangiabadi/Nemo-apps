#!/usr/bin/env python3
"""
Create a NEMO PDF invoice from an edited Excel invoice workbook.

This module is intentionally small: it reads the "Invoice" sheet produced by
``nemo_invoice_generator_with_pdf.py``, reconstructs the invoice rows, and then
uses the existing ReportLab PDF renderer so the output format stays consistent.
"""

from __future__ import annotations

import datetime as dt
import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

import nemo_invoice_generator_with_pdf as invoice_logic


DETAIL_LABS = set(invoice_logic.DESIRED_LAB_ORDER)
SKIPPED_SECTION_TITLES = {"Access fee", "Project fees summary"}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _money_value(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    if cleaned in {"", "-"}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _quantity_value(value: object) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _period_from_month_label(value: object) -> str:
    label = _cell_text(value)
    for fmt in ("%b%Y", "%B%Y", "%b %Y", "%B %Y", "%Y-%m"):
        try:
            parsed = dt.datetime.strptime(label, fmt)
            return parsed.strftime("%Y-%m")
        except ValueError:
            pass
    return label or dt.datetime.now().strftime("%Y-%m")


def _start_datetime(value: object) -> Optional[dt.datetime]:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min)
    return invoice_logic.parse_nemo_datetime(value)


def _section_title(value: object) -> str:
    return _cell_text(value)


def _is_detail_section_title(value: object) -> bool:
    return _section_title(value) in DETAIL_LABS


def _read_project_types(ws) -> dict[str, str]:
    project_types: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        if _section_title(ws.cell(row, 1).value) != "Project fees summary":
            continue

        header_row = row + 1
        headers = {
            _cell_text(ws.cell(header_row, col).value): col
            for col in range(1, ws.max_column + 1)
        }
        project_col = headers.get("Project")
        type_col = headers.get("Project Type")
        if not project_col or not type_col:
            return project_types

        current = header_row + 1
        while current <= ws.max_row:
            first_cell = _cell_text(ws.cell(current, 1).value)
            if first_cell in {
                "Usage charges total",
                "Access fee",
                "Invoice total",
            }:
                break
            project = _cell_text(ws.cell(current, project_col).value)
            project_type = _cell_text(ws.cell(current, type_col).value)
            if project and project_type:
                project_types[project] = project_type
            current += 1
        return project_types
    return project_types


def _read_access_fee(ws) -> float:
    for row in range(1, ws.max_row + 1):
        if _cell_text(ws.cell(row, 1).value).lower() == "access fee":
            value = ws.cell(row, 2).value
            if value not in (None, ""):
                return _money_value(value)
    return 0.0


def _detail_rows_from_sheet(ws, project_types: dict[str, str]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    row = 1

    while row <= ws.max_row:
        lab = _section_title(ws.cell(row, 1).value)
        if lab not in DETAIL_LABS:
            row += 1
            continue

        header_row = row + 1
        headers = {
            _cell_text(ws.cell(header_row, col).value): col
            for col in range(1, ws.max_column + 1)
        }
        required = ["Date", "User", "Description", "Type", "Project", "Quantity", "Rate", "Cost"]
        if any(name not in headers for name in required):
            row += 1
            continue

        current = header_row + 1
        while current <= ws.max_row:
            first = _section_title(ws.cell(current, 1).value)
            if (
                _is_detail_section_title(first)
                or first in SKIPPED_SECTION_TITLES
                or first == "Project fees summary"
            ):
                break
            if any(
                _cell_text(ws.cell(current, col).value) == "Subtotal"
                for col in range(1, ws.max_column + 1)
            ):
                current += 1
                break

            project = _cell_text(ws.cell(current, headers["Project"]).value)
            description = _cell_text(ws.cell(current, headers["Description"]).value)
            user = _cell_text(ws.cell(current, headers["User"]).value)
            if not any((project, description, user)):
                current += 1
                continue

            row_data = {
                "Type": _cell_text(ws.cell(current, headers["Type"]).value),
                "User": user,
                "Item": description,
                "Item_norm": description,
                "Project": project,
                "Application identifier": project_types.get(project, ""),
                "Start_dt": _start_datetime(ws.cell(current, headers["Date"]).value),
                "Rate": _cell_text(ws.cell(current, headers["Rate"]).value),
                "Cost": _money_value(ws.cell(current, headers["Cost"]).value),
                "Quantity": _quantity_value(ws.cell(current, headers["Quantity"]).value),
                "Subsidy": _money_value(ws.cell(current, headers.get("Subsidy", 0)).value)
                if headers.get("Subsidy")
                else 0.0,
                "Lab": lab,
                "IsConsumable": lab == "Consumable",
            }
            rows.append(row_data)
            current += 1

        row = current

    return rows


def parse_excel_invoice(workbook_path: str | os.PathLike[str]) -> tuple[pd.DataFrame, dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        ws = workbook["Invoice"] if "Invoice" in workbook.sheetnames else workbook.active
        project_types = _read_project_types(ws)
        rows = _detail_rows_from_sheet(ws, project_types)
        if not rows:
            raise ValueError(
                "No invoice line items were found. Upload an Excel invoice generated by the NEMO Invoice Generator."
            )

        period = _period_from_month_label(ws["E4"].value)
        df = pd.DataFrame(rows)
        df["Period"] = period
        df["Application identifier"] = (
            df["Application identifier"].fillna("").astype(str).str.strip()
        )
        if (df["Application identifier"] == "").any():
            df.loc[df["Application identifier"] == "", "Application identifier"] = "Local"

        metadata = {
            "pi_display_name": _cell_text(ws["B4"].value) or "UNKNOWN_PI",
            "pi_email": "" if _cell_text(ws["B5"].value) == "N/A" else _cell_text(ws["B5"].value),
            "period_ym": period,
            "invoice_number": _cell_text(ws["K4"].value)
            or invoice_logic.make_invoice_number(period),
            "internal_fee": _read_access_fee(ws),
        }
        return df, metadata
    finally:
        workbook.close()


def convert_excel_invoice_to_pdf(
    workbook_path: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    *,
    pi_email: str = "",
    logo_path: Optional[str] = None,
) -> str:
    if not invoice_logic._pdf_available():
        raise RuntimeError(
            "PDF generation requested, but reportlab is not installed in this Python environment."
        )

    df, metadata = parse_excel_invoice(workbook_path)
    output_path = (
        Path(output_dir)
        / f"{invoice_logic.safe_filename(str(metadata['pi_display_name']))} "
        f"{invoice_logic.month_label(str(metadata['period_ym']))}.pdf"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    invoice_logic.create_invoice_pdf(
        df,
        pi_display_name=str(metadata["pi_display_name"]),
        pi_email=pi_email or str(metadata["pi_email"]),
        period_ym=str(metadata["period_ym"]),
        invoice_number=str(metadata["invoice_number"]),
        pdf_path=str(output_path),
        logo_path=logo_path,
        internal_fee_override=float(metadata["internal_fee"]),
    )
    return str(output_path)
