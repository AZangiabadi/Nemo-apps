#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import nemo_invoice_generator_with_pdf as inv


APRIL_CSV = Path("/Users/nemohub/Downloads/usage_export_April_06_17_2026-11_25_58.csv")
MAY_CSV = Path("/Users/nemohub/Downloads/usage_export_May_06_17_2026-11_25_31.csv")
OUTPUT_DIR = Path("/Users/nemohub/Nemo-apps/outputs/usage_cap_report")
OUTPUT_XLSX = OUTPUT_DIR / "NEMO_April_May_2026_Cap_Analysis_with_Staff.xlsx"


MONEY_COLUMNS = {
    "No Cap Usage Charges",
    "Hourly-Capped Usage Charges",
    "Final Usage Charges",
    "No-Hourly Project-Capped Usage Charges",
    "Access Fees",
    "No Cap Total Income",
    "Hourly-Capped Total Income",
    "Final Total Income",
    "No-Hourly Project-Capped Total Income",
    "Staff Charges Already Included",
    "Additional Staff-Application Charges",
    "Final Total Income Including Staff Application",
    "No Cap Total Income Including Staff Application",
    "Hourly Cap Reduction",
    "Project Cap Reduction",
    "Total Cap Reduction",
    "Usage Before Project Cap",
    "Capped Billable Usage",
    "Project Cap Savings",
    "No-Hourly Usage Before Project Cap",
    "No-Hourly Capped Billable Usage",
    "No-Hourly Project Cap Savings",
    "Original Cost",
    "Hourly-Capped Cost",
    "Reduction",
    "Cost",
    "Final Cost",
}


def parse_period(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(
        series.astype(str).str.replace(" @ ", " ", regex=False),
        format="%m/%d/%Y %I:%M %p",
        errors="coerce",
    )
    return parsed.dt.strftime("%Y-%m").fillna("Unknown")


def prepare_base(csv_path: Path, source_month: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df["Source Month"] = source_month
    df["Application identifier"] = df["Application identifier"].astype(str).str.strip()
    df = df[df["Application identifier"].isin(inv.INVOICE_APPLICATION_IDENTIFIERS)].copy()

    df["Start_dt"] = df["Start time"].apply(inv.parse_nemo_datetime)
    df["End_dt"] = df["End time"].apply(inv.parse_nemo_datetime)
    df["Item_norm"] = df["Item"].apply(inv.normalize_item)
    df["IsConsumable"] = df["Type"].apply(inv._is_consumable_type)
    df["IsMissedReservation"] = df.apply(inv._row_contains_missed_reservation_text, axis=1)
    df["IsStaffCharge"] = df.apply(inv._row_is_staff_charge, axis=1)
    df["IsToolUsageCharge"] = ~df["IsConsumable"] & ~df["IsMissedReservation"]
    df["Lab"] = df["Item_norm"].map(inv.TOOL_TO_LAB).fillna("Unmapped")
    df["Lab"] = df["Lab"].map(inv.LAB_NAME_MAP).fillna(df["Lab"])
    df = inv.apply_staff_time_lab_associations(df)
    df["Cost"] = pd.to_numeric(df["Cost"], errors="coerce").fillna(0.0).astype(float)
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").astype(float)
    df["Period"] = df["Start_dt"].apply(inv.period_from_start_dt)
    df["Billable User Key"] = df.apply(inv.resolve_billable_user_key, axis=1)
    df["PI_key"] = df["Project"].apply(inv.extract_pi_code_from_project)
    return df[~df["IsMissedReservation"].fillna(False)].reset_index(drop=True)


def prepare_raw(paths: list[tuple[Path, str]]) -> pd.DataFrame:
    frames = []
    for csv_path, source_month in paths:
        df = pd.read_csv(csv_path)
        df["Source Month"] = source_month
        df["Cost"] = pd.to_numeric(df["Cost"], errors="coerce").fillna(0.0).astype(float)
        df["Period"] = parse_period(df["Start time"])
        df["Item_norm"] = df["Item"].apply(inv.normalize_item)
        df["IsStaffCharge"] = df.apply(inv._row_is_staff_charge, axis=1)
        frames.append(df)
    return pd.concat(frames, ignore_index=True)


def staff_breakdown(raw: pd.DataFrame, invoice_base: pd.DataFrame) -> pd.DataFrame:
    invoice_staff = (
        invoice_base[invoice_base["IsStaffCharge"].fillna(False)]
        .groupby("Period", dropna=False)["Cost"]
        .sum()
        .rename("Staff Charges Already Included")
    )
    additional_staff_app = (
        raw[
            raw["Application identifier"].astype(str).str.strip().eq("Staff")
            & ~raw["Type"].astype(str).str.lower().eq("missed_reservation")
        ]
        .groupby("Period", dropna=False)["Cost"]
        .sum()
        .rename("Additional Staff-Application Charges")
    )
    staff_time_item = (
        raw[
            raw["Item_norm"].astype(str).str.lower().eq("staff time")
            & ~raw["Type"].astype(str).str.lower().eq("missed_reservation")
        ]
        .groupby("Period", dropna=False)["Cost"]
        .sum()
        .rename("Raw Item=Staff Time Charges")
    )
    out = pd.concat([invoice_staff, additional_staff_app, staff_time_item], axis=1).fillna(0.0)
    out["Total Visible Staff-Related Charges"] = (
        out["Staff Charges Already Included"] + out["Additional Staff-Application Charges"]
    )
    return out.reset_index().rename(columns={"index": "Period"})


def staff_detail(raw: pd.DataFrame, invoice_base: pd.DataFrame) -> pd.DataFrame:
    raw_staff_app = raw[
        raw["Application identifier"].astype(str).str.strip().eq("Staff")
        & ~raw["Type"].astype(str).str.lower().eq("missed_reservation")
    ].copy()
    raw_staff_app["Staff Bucket"] = "Additional Staff-Application Charges"

    included = invoice_base[invoice_base["IsStaffCharge"].fillna(False)].copy()
    included["Staff Bucket"] = "Staff Charges Already Included"

    cols = [
        "Staff Bucket",
        "Period",
        "Source Month",
        "Type",
        "User",
        "Username",
        "Item",
        "Application identifier",
        "Project",
        "Start time",
        "End time",
        "Quantity",
        "Rate",
        "Cost",
    ]
    return pd.concat([included[cols], raw_staff_app[cols]], ignore_index=True).sort_values(
        ["Staff Bucket", "Period", "Cost"], ascending=[True, True, False], kind="stable"
    )


def access_fees(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (period, pi_key), grp in df.groupby(["Period", "PI_key"], dropna=False):
        fee = inv.internal_facility_fee_for_group(grp)
        if fee > 0:
            rows.append({"Period": period, "PI/User Group": pi_key, "Access Fees": fee})
    return pd.DataFrame(rows)


def project_cap_hits(df_before_project_cap: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    columns = [
        "Period",
        "User Key",
        "User",
        "Project",
        "Application identifier",
        f"{prefix}Project Cap",
        f"{prefix}Usage Before Project Cap",
        f"{prefix}Capped Billable Usage",
        f"{prefix}Project Cap Savings",
        "Tool Usage Rows",
        "Top Instruments",
    ]
    rows = []
    if df_before_project_cap.empty:
        return pd.DataFrame(columns=columns)

    for (period, user_key, project, app_id), grp in df_before_project_cap.groupby(
        ["Period", "Billable User Key", "Project", "Application identifier"], dropna=False
    ):
        cap = inv.PROJECT_CHARGE_CAP_BY_APPLICATION.get(str(app_id))
        if cap is None:
            continue
        tool_grp = grp[
            grp["IsToolUsageCharge"].fillna(False).astype(bool)
            & ~grp["IsStaffCharge"].fillna(False).astype(bool)
        ]
        original = float(tool_grp["Cost"].fillna(0).sum())
        if original <= cap:
            continue
        rows.append(
            {
                "Period": period,
                "User Key": user_key,
                "User": ", ".join(sorted({str(x) for x in tool_grp["User"].dropna().unique()})),
                "Project": project,
                "Application identifier": app_id,
                f"{prefix}Project Cap": cap,
                f"{prefix}Usage Before Project Cap": original,
                f"{prefix}Capped Billable Usage": cap,
                f"{prefix}Project Cap Savings": original - cap,
                "Tool Usage Rows": int(tool_grp.shape[0]),
                "Top Instruments": ", ".join(
                    tool_grp.groupby("Item_norm")["Cost"].sum().sort_values(ascending=False).head(3).index.astype(str)
                ),
            }
        )
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).sort_values(
        [f"{prefix}Project Cap Savings", f"{prefix}Usage Before Project Cap"],
        ascending=[False, False],
    )


def scenario_summary(
    base: pd.DataFrame,
    hourly: pd.DataFrame,
    final: pd.DataFrame,
    no_hourly_project: pd.DataFrame,
    staff: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    fee_by_period = access_fees(base).groupby("Period")["Access Fees"].sum()
    staff_by_period = staff.set_index("Period")
    periods = sorted(set(base["Period"].dropna()) | set(staff["Period"].dropna()))
    for period in periods:
        base_p = base[base["Period"] == period]
        hourly_p = hourly[hourly["Period"] == period]
        final_p = final[final["Period"] == period]
        no_hourly_project_p = no_hourly_project[no_hourly_project["Period"] == period]
        fees = float(fee_by_period.get(period, 0.0))
        staff_included = float(staff_by_period.get("Staff Charges Already Included", pd.Series()).get(period, 0.0))
        staff_additional = float(staff_by_period.get("Additional Staff-Application Charges", pd.Series()).get(period, 0.0))
        no_cap_usage = float(base_p["Cost"].fillna(0).sum())
        hourly_usage = float(hourly_p["Cost"].fillna(0).sum())
        final_usage = float(final_p["Cost"].fillna(0).sum())
        no_hourly_project_usage = float(no_hourly_project_p["Cost"].fillna(0).sum())
        rows.append(
            {
                "Period": period,
                "No Cap Usage Charges": no_cap_usage,
                "Hourly-Capped Usage Charges": hourly_usage,
                "Final Usage Charges": final_usage,
                "No-Hourly Project-Capped Usage Charges": no_hourly_project_usage,
                "Access Fees": fees,
                "Staff Charges Already Included": staff_included,
                "Additional Staff-Application Charges": staff_additional,
                "No Cap Total Income": no_cap_usage + fees,
                "No Cap Total Income Including Staff Application": no_cap_usage + fees + staff_additional,
                "Final Total Income": final_usage + fees,
                "Final Total Income Including Staff Application": final_usage + fees + staff_additional,
                "No-Hourly Project-Capped Total Income": no_hourly_project_usage + fees,
                "Hourly Cap Reduction": no_cap_usage - hourly_usage,
                "Project Cap Reduction": hourly_usage - final_usage,
                "Total Cap Reduction": no_cap_usage - final_usage,
                "Users Hitting Hourly Cap": int(hourly_p.loc[hourly_p["Original Cost"].notna(), "Billable User Key"].nunique()),
                "User/Project Groups Hitting Project Cap": int(project_cap_hits(hourly_p).shape[0]),
                "User/Project Groups Hitting Project Cap Without Hourly Cap": int(project_cap_hits(base_p, prefix="No-Hourly ").shape[0]),
            }
        )
    total = pd.DataFrame(rows)
    total_row = {"Period": "Total"}
    for col in total.columns:
        if col != "Period":
            total_row[col] = total[col].sum()
    return pd.concat([total, pd.DataFrame([total_row])], ignore_index=True)


def hourly_cap_hits(hourly: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    hit = hourly[hourly["Original Cost"].notna()].copy()
    if hit.empty:
        return pd.DataFrame(), pd.DataFrame()
    hit["Original Hours"] = hit["Original Quantity"].astype(float) / 60.0
    hit["Billable Hours After Cap"] = hit["Quantity"].astype(float) / 60.0
    hit["Reduction"] = hit["Original Cost"].astype(float) - hit["Cost"].astype(float)
    hit = hit.rename(columns={"Item_norm": "Instrument", "Cost": "Hourly-Capped Cost"})
    details = hit[
        [
            "Period",
            "User",
            "Username",
            "Instrument",
            "Project",
            "Application identifier",
            "Start time",
            "End time",
            "Max Billable Hours",
            "Original Hours",
            "Billable Hours After Cap",
            "Original Cost",
            "Hourly-Capped Cost",
            "Reduction",
        ]
    ].sort_values(["Reduction", "Original Cost"], ascending=[False, False])
    top = (
        hit.groupby(["User", "Username", "Instrument", "Application identifier"], dropna=False)
        .agg(
            **{
                "Capped Sessions": ("Reduction", "size"),
                "Original Hours": ("Original Hours", "sum"),
                "Billable Hours After Cap": ("Billable Hours After Cap", "sum"),
                "Original Cost": ("Original Cost", "sum"),
                "Hourly-Capped Cost": ("Hourly-Capped Cost", "sum"),
                "Reduction": ("Reduction", "sum"),
            }
        )
        .reset_index()
        .sort_values(["Reduction", "Capped Sessions"], ascending=[False, False])
        .head(30)
    )
    return top, details


def lab_summary(df: pd.DataFrame, scenario: str) -> pd.DataFrame:
    out = df.groupby(["Period", "Lab"], dropna=False)["Cost"].sum().reset_index()
    out.insert(0, "Scenario", scenario)
    return out


def write_df(ws, df: pd.DataFrame, start_row: int = 1, table_name: str | None = None) -> None:
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(start_row, col_idx, col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            column_name = str(df.columns[col_idx - 1])
            if column_name in MONEY_COLUMNS or "Savings" in column_name or "Reduction" in column_name:
                cell.number_format = '$#,##0.00;[Red]($#,##0.00)'
            elif "Hours" in column_name:
                cell.number_format = "0.00"
    last_row = start_row + max(len(df), 1)
    last_col = max(len(df.columns), 1)
    ws.freeze_panes = ws.cell(start_row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    adjust_widths(ws)


def adjust_widths(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 11), 55)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    thin = Side(style="thin", color="D9E2F3")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
    wb.save(path)


def cap_report_title(periods: list[str]) -> str:
    known_periods = [period for period in periods if period and period != "Unknown"]
    if not known_periods:
        return "NEMO Usage Cap Analysis"
    if len(known_periods) == 1:
        return f"NEMO Usage Cap Analysis - {known_periods[0]}"
    return f"NEMO Usage Cap Analysis - {known_periods[0]} through {known_periods[-1]}"


def build_usage_cap_report(
    paths: list[tuple[Path, str]],
    output_path: Path,
) -> dict[str, object]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    raw = prepare_raw(paths)
    base = pd.concat([prepare_base(path, label) for path, label in paths], ignore_index=True)
    if base.empty:
        raise ValueError("No invoice-compatible usage rows were found in the uploaded CSV file(s).")

    hourly = inv.apply_max_session_charge_caps(base.copy())
    final = inv.apply_project_charge_caps(hourly.copy())
    no_hourly_project = inv.apply_project_charge_caps(base.copy())
    staff = staff_breakdown(raw, base)

    summary = scenario_summary(base, hourly, final, no_hourly_project, staff)
    top_hourly, hourly_details = hourly_cap_hits(hourly)
    project_hits = project_cap_hits(hourly)
    no_hourly_project_hits = project_cap_hits(base, prefix="No-Hourly ")
    lab = pd.concat(
        [
            lab_summary(base, "No Caps"),
            lab_summary(hourly, "Hourly Cap Only"),
            lab_summary(final, "Hourly + Project Caps"),
            lab_summary(no_hourly_project, "Project Cap Only, No Hourly Cap"),
        ],
        ignore_index=True,
    )
    source_period_audit = (
        base.groupby(["Source Month", "Period"], dropna=False)
        .agg(Rows=("Cost", "size"), Cost=("Cost", "sum"))
        .reset_index()
        .sort_values(["Source Month", "Period"], kind="stable")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    periods = sorted(set(base["Period"].dropna()) | set(staff["Period"].dropna()))
    ws["A1"] = cap_report_title(periods)
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws["A2"] = f"Generated {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = (
        "Staff charges already included are supported-application staff charges inside the invoice-compatible totals. "
        "Additional Staff-application charges are shown separately because the invoice code filters that application identifier out."
    )
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:K3")
    write_df(ws, summary, start_row=5, table_name="ExecutiveSummary")

    chart = BarChart()
    chart.title = "Final Income With/Without Staff Application"
    chart.y_axis.title = "USD"
    data = Reference(ws, min_col=11, max_col=12, min_row=5, max_row=5 + len(summary) - 1)
    cats = Reference(ws, min_col=1, min_row=6, max_row=5 + len(summary) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "U5")

    sheets = [
        ("Staff Charges", staff, "StaffCharges"),
        ("Staff Charge Detail", staff_detail(raw, base), "StaffChargeDetail"),
        ("Monthly Lab Summary", lab, "MonthlyLabSummary"),
        ("Top 30 Hourly Cap", top_hourly, "TopHourlyCap"),
        ("Hourly Cap Details", hourly_details, "HourlyCapDetails"),
        ("Project Cap Hits", project_hits, "ProjectCapHits"),
        ("No Hourly Project Cap", no_hourly_project_hits, "NoHourlyProjectCap"),
        ("Access Fees", access_fees(base), "AccessFees"),
        ("Source Period Audit", source_period_audit, "SourcePeriodAudit"),
    ]
    for name, df, table_name in sheets:
        sheet = wb.create_sheet(name)
        write_df(sheet, df, table_name=table_name)

    assumptions = wb.create_sheet("Assumptions")
    assumptions_df = pd.DataFrame(
        [
            ["Source files", str(APRIL_CSV), ""],
            ["Source files", str(MAY_CSV), ""],
            ["Application filter", ", ".join(inv.INVOICE_APPLICATION_IDENTIFIERS), "Matches invoice generator for invoice-compatible totals."],
            ["Additional staff application", "Application identifier = Staff", "Shown separately and added only in the explicit including-staff-application columns."],
            ["Staff already included", "Rows marked staff_charge or Item = Staff Time in supported application identifiers", "Already part of invoice-compatible usage totals."],
            ["Excluded rows", "Missed reservations", "Invoice generation removes missed reservations from output."],
            ["Hourly cap order", "Hourly/session caps applied before project caps", "Matches load_and_prepare()."],
            ["Project cap grouping", "Period + Billable User Key + Project + Application identifier", "Matches apply_project_charge_caps(). Staff charges are excluded from project cap reduction."],
            ["Access fee", "Per PI/month invoice group, highest application fee in group", "Included in total income, not reduced by caps."],
            ["PI grouping", "Last token of Project string", "No NEMO API token was used."],
            ["Period handling", "Grouped by parsed Start time", "Some source-file rows cross month boundaries."],
        ],
        columns=["Topic", "Value", "Notes"],
    )
    write_df(assumptions, assumptions_df, table_name="Assumptions")

    wb.save(output_path)
    wb.close()
    style_workbook(output_path)
    total_row = summary[summary["Period"] == "Total"].iloc[0].to_dict()
    return {
        "output_path": str(output_path),
        "periods": periods,
        "source_file_count": len(paths),
        "row_count": int(base.shape[0]),
        "final_total_income": float(total_row.get("Final Total Income", 0.0) or 0.0),
        "final_total_income_including_staff": float(
            total_row.get("Final Total Income Including Staff Application", 0.0) or 0.0
        ),
        "additional_staff_application_charges": float(
            total_row.get("Additional Staff-Application Charges", 0.0) or 0.0
        ),
        "hourly_cap_reduction": float(total_row.get("Hourly Cap Reduction", 0.0) or 0.0),
        "project_cap_reduction": float(total_row.get("Project Cap Reduction", 0.0) or 0.0),
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    paths = [(APRIL_CSV, "April"), (MAY_CSV, "May")]
    result = build_usage_cap_report(paths, OUTPUT_XLSX)
    print(result["output_path"])
    raw = prepare_raw(paths)
    base = pd.concat([prepare_base(path, label) for path, label in paths], ignore_index=True)
    hourly = inv.apply_max_session_charge_caps(base.copy())
    final = inv.apply_project_charge_caps(hourly.copy())
    no_hourly_project = inv.apply_project_charge_caps(base.copy())
    staff = staff_breakdown(raw, base)
    summary = scenario_summary(base, hourly, final, no_hourly_project, staff)
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
