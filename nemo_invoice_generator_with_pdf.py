#!/usr/bin/env python3
"""
Generate per-PI (and per-month) Excel invoices from a NEMO "usage export" CSV,
AND (optionally) generate matching PDF invoices.

What it does
------------
- Reads the NEMO usage export CSV (columns like Type/User/Item/Project/Rate/Cost/...).
- Keeps ONLY rows where Application identifier is one of:
  "Local", "CDG", "Industry", or "External Academia".
- Splits charges into labs using your tool->lab mapping.
- Groups rows into one invoice per PI per month.
- Writes an "Invoice" sheet in an XLSX with:
    * Summary totals (cost per lab + access fee + total)
    * Separate tables per lab (Cleanroom, SMCL, Electron Microscopy Lab, plus "Consumable")
    * A "Project fees summary" table at the bottom (total charges per project, broken down by lab)
- Writes a PDF invoice alongside each XLSX (invoice-style layout, with optional logo top-left).

PI grouping
-----------
Best: provide a NEMO API token so the script can map Project -> contact_name/contact_email.
Fallback (no token): extract a PI code from the last token of the Project string (often a UNI like ML3745).

Run (GUI)
---------
    python nemo_invoice_generator_with_pdf.py

- Picks the CSV using a file dialog.
- Prompts for your API token (optional; leave blank to skip).
- (If generating PDFs) optionally asks you to select a logo image (PNG/JPG).

Run (no GUI / servers)
----------------------
    python nemo_invoice_generator_with_pdf.py --csv /path/to/usage_export.csv --no-gui

If you want a logo in no-GUI mode:
    python nemo_invoice_generator_with_pdf.py --csv ... --no-gui --logo /path/to/columbia_logo.png

Dependencies
------------
- pandas
- openpyxl
- requests (only if using API token)
- reportlab (only if generating PDFs)
"""

from __future__ import annotations

import argparse
import datetime as dt
import getpass
import math
import os
import re
import sys
import time
import traceback
import zipfile
from dataclasses import dataclass
from typing import Callable, Dict, Iterable, List, Optional, Tuple
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# Optional: requests for NEMO API access
try:
    import requests  # type: ignore
except Exception:
    requests = None

# Optional: reportlab for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.units import inch
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )
    from xml.sax.saxutils import escape as _xml_escape
except Exception:
    # We'll handle missing reportlab gracefully (PDF generation will be disabled)
    colors = None
    letter = None
    landscape = None
    ParagraphStyle = None
    getSampleStyleSheet = None
    inch = None
    ImageReader = None
    Image = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None
    Table = None
    TableStyle = None
    PageBreak = None
    _xml_escape = None


# -----------------------------
# Local configuration
# -----------------------------
NEMO_BASE_URL = "https://nemo.cni.columbia.edu"
INVOICE_TIMEZONE = ZoneInfo("America/New_York")

# Internal facility fee (USD) by account/application identifier
INTERNAL_FACILITY_FEE_BY_APPLICATION = {
    "Local": 50.0,
    "CDG": 50.0,
    "External Academia": 75.0,
    "Industry": 150.0,
}

PROJECT_CHARGE_CAP_BY_APPLICATION = {
    "Local": 1500.0,
    "CDG": 1500.0,
    "External Academia": 2500.0,
    "Industry": 4500.0,
}

NEMO_METADATA_CACHE_TTL_SECONDS = max(
    0, int(os.environ.get("NEMO_INVOICE_METADATA_CACHE_SECONDS", "21600"))
)
NEMO_METADATA_CACHE: Dict[tuple[str, str, str], tuple[float, dict]] = {}

# Application identifiers included in invoice generation
INVOICE_APPLICATION_IDENTIFIERS = ("Local", "CDG", "Industry", "External Academia")


def invoice_generated_at() -> dt.datetime:
    """Return the invoice generation timestamp in Eastern time for Excel/PDF headers."""
    return dt.datetime.now(INVOICE_TIMEZONE)

# Subsidy applied to CDG usage items for reporting only.
# We do not modify the costs from CSV; instead we record theoretical savings.
# For CDG, the proxy subsidy is 1/9 of the charged cost (consistent with a 10% effective discount).


# -----------------------------
# Tool -> Lab mapping (provided)
# -----------------------------
TOOL_LAB_CSV = """CMP,Cleanroom
Dicing Saw,Cleanroom
TPT Wirebonder HB10,Cleanroom
AJA Orion-3 Metal Sputtering System,Cleanroom
AJA Orion-8 Dielectrics Sputtering System,Cleanroom
Angstrom High Vacuum,Cleanroom
Angstrom Metals Deposition System,Cleanroom
Cambridge NanoTech ALD,Cleanroom
Cressington Sputter Coater,Cleanroom
Edwards Thermal Evaporator 1,Cleanroom
Oxford PECVD,Cleanroom
Parylene Coater,Cleanroom
Solaris RTA,Cleanroom
Across TF1700,Cleanroom
Anatech Plasma Asher,Cleanroom
Diener Plasma Etch,Cleanroom
Oxford ICP-DRIE F-based Cobra300,Cleanroom
Oxford ICP-RIE Cl based Cobra III-V,Cleanroom
Oxford ICP RIE - direct load,Cleanroom
UVOCs UV Ozone Cleaner,Cleanroom
Beamer,Cleanroom
Elionix BODEN 50f EBL,Cleanroom
KLA P17 Profiler,Cleanroom
Lakeshore Hall System,Cleanroom
Nanomagnetics ezAFM,Cleanroom
NovaNano SEM,Cleanroom
Park AFM,Cleanroom
Woollam Alpha-SE ellipsometer,Cleanroom
Wyko NT9100 Optical Profiler,Cleanroom
BlueM Oven,Cleanroom
Heidelberg (3um) Laser Writer,Cleanroom
Heidelberg DWL 66+ Laser Writer,Cleanroom
Laurell Spinner 1,Cleanroom
Litho Hood 1 Spinner 1,Cleanroom
Litho Hood 1 Spinner 2,Cleanroom
Litho Hood 2,Cleanroom
Litho Hood 3 Spinner 3,Cleanroom
Litho - Solvent Tank,Cleanroom
Suss MA6 DUV Mask Aligner,Cleanroom
Suss MA6 Mask Aligner,Cleanroom
Vacuum Oven,Cleanroom
YES (HMDS) Oven,Cleanroom
General Acids Hood,Cleanroom
General Base Hood,Cleanroom
RCA Station,Cleanroom
ASTAR Analysis PC,Electron Microscopy
EBSD Analysis PC,Electron Microscopy
FEI Talos F200x S-TEM,Electron Microscopy
Light Zeiss Microscope,Electron Microscopy
ZEISS SEM,Electron Microscopy
Diamond Saw,Electron Microscopy
Dimple Grinder,Electron Microscopy
FIB Sample Preparation,Electron Microscopy
Grinder-Polisher,Electron Microscopy
Microtome,Electron Microscopy
PIPS II,Electron Microscopy
Plasma Cleaner,Electron Microscopy
TEM BIO Samples,Electron Microscopy
Agilent 1260 Infinity GPC,SMCL
Agilent 8453 UV-Vis Spectrophotometer,SMCL
Agilent SuperNova SCXRD,SMCL
Autofinder 1,SMCL
Autofinder 2,SMCL
Bal-Tec CPD,SMCL
Bruker Dimensions FastScan AFM,SMCL
Horiba XploRA micro-Raman,SMCL
Malvern Zetasizer Nano-ZS,SMCL
Micrometrics ASAP 2020 HV BET analyzer,SMCL
PANalytical XPert3 Powder XRD,SMCL
Phi 5500 XPS,SMCL
Renishaw inVia micro-Raman,SMCL
Rigaku SmartLab XRD,SMCL
Rigaku XtaLAB Synergy-S SCXRD,SMCL
TA Instruments Q500 TGA,SMCL
Tosoh EcoSEC RI-UV GPC,SMCL
Woollam Variable Angle Ellipsometer,SMCL
"""


def build_tool_to_lab_map(tool_lab_csv: str = TOOL_LAB_CSV) -> Dict[str, str]:
    tool_to_lab: Dict[str, str] = {}
    for line in tool_lab_csv.strip().splitlines():
        tool, lab = [x.strip() for x in line.split(",", 1)]
        tool_to_lab[tool] = lab
    return tool_to_lab


TOOL_TO_LAB = build_tool_to_lab_map()
LAB_NAME_MAP = {
    "Cleanroom": "Cleanroom",
    "SMCL": "SMCL",
    "Electron Microscopy": "Electron Microscopy Lab",
    "Unmapped": "Consumable",  # renamed per your request
}

DESIRED_LAB_ORDER = ["Cleanroom", "SMCL", "Electron Microscopy Lab", "Consumable"]

# Max billable hours per session, sourced from "Tool Rates.xlsx" Sheet2.
TOOL_MAX_HOURS_BY_TOOL_ID = {
    2: 3.0,
    3: 3.0,
    4: 3.0,
    5: 3.0,
    6: 3.0,
    7: 3.0,
    8: 8.0,
    9: 10.0,
    10: 4.0,
    11: 4.0,
    14: 4.0,
    15: 8.0,
    16: 6.0,
    18: 8.0,
    19: 8.0,
    20: 2.0,
    21: 4.0,
    22: 3.0,
    23: 9.0,
    24: 9.0,
    26: 4.0,
    27: 4.0,
    28: 4.0,
    29: 6.0,
    30: 6.0,
    31: 4.0,
    32: 6.0,
    33: 4.0,
    34: 4.0,
    35: 4.0,
    36: 4.0,
    37: 4.0,
    38: 4.0,
    39: 4.0,
    40: 3.0,
    41: 3.0,
    42: 3.0,
    43: 2.0,
    44: 4.0,
    45: 6.0,
    46: 6.0,
    47: 2.0,
    48: 3.0,
    49: 8.0,
    51: 5.0,
    52: 8.0,
    53: 9.0,
    54: 5.0,
    55: 10.0,
    56: 10.0,
    57: 8.0,
    59: 9.0,
    60: 10.0,
    61: 12.0,
    62: 10.0,
    63: 10.0,
    64: 10.0,
    65: 10.0,
    66: 12.0,
    67: 8.0,
    68: 8.0,
    69: 4.0,
    70: 4.0,
    71: 4.0,
    72: 4.0,
    73: 4.0,
    74: 4.0,
    77: 12.0,
    78: 12.0,
}

TOOL_MAX_HOURS_BY_NAME = {
    "laurellspinner": 3.0,
    "lithohood1spinner1": 3.0,
    "lithohood1spinner2": 3.0,
    "lithohood2": 3.0,
    "lithohood3spinner3": 3.0,
    "lithosolventtank": 3.0,
    "heidelberg3micronlaserwriter": 8.0,
    "heidelbergdw66laserwriter": 10.0,
    "duvma6maskaligner": 4.0,
    "ma6maskaligner": 4.0,
    "yeshmdsoven": 4.0,
    "feitalostem": 8.0,
    "zeisssigmasem": 6.0,
    "elionixboden50febl": 8.0,
    "beamer": 8.0,
    "edwardsthermalevaporator": 2.0,
    "lightzeissmicroscope": 4.0,
    "criticalpointdryer": 3.0,
    "autofinder1": 9.0,
    "autofinder2": 9.0,
    "oxfordpecvd": 4.0,
    "angstromevovacsystem": 4.0,
    "angstromhighvacuumevaporator": 4.0,
    "ajadielectricsputter": 6.0,
    "ajametalsputter": 6.0,
    "cambridgenanotechald": 4.0,
    "parylenecoater": 6.0,
    "solarisrta": 4.0,
    "acrosstf1700": 4.0,
    "oxfordicprieclbasedcobraiiiv": 4.0,
    "oxfordicpdriefbasedcobra300": 4.0,
    "oxfordicpriedirectload": 4.0,
    "dienerplasmaetch": 4.0,
    "anatechplasmaasher": 4.0,
    "rcastation": 3.0,
    "generalacidshood": 3.0,
    "generalbasehood": 3.0,
    "uvozone": 2.0,
    "tptwirebonder": 4.0,
    "dicingsaw": 6.0,
    "chemicalmechanicalpolishing": 6.0,
    "klaprofilometer": 2.0,
    "wykont9100opticalprofiler": 3.0,
    "novananosem": 8.0,
    "nanomagneticsezafm": 5.0,
    "parkafm": 8.0,
    "bet": 9.0,
    "agilentecosecgpc": 5.0,
    "agilent8453uvvisspectrophotometer": 10.0,
    "agilentsupernovascxrd": 10.0,
    "brukerdimensionfastscanafm": 8.0,
    "horibamicroraman": 9.0,
    "malvernzetasizernanozs": 10.0,
    "panalyticalxpert3powderxrd": 12.0,
    "phixps": 10.0,
    "renishawinviamicroraman": 10.0,
    "smartlabxrd": 10.0,
    "synergysscxrd": 10.0,
    "tainstrumentsq500tga": 12.0,
    "woollamvariableangleellipsometer": 8.0,
    "woollamalphaseellipsometer": 8.0,
    "pipsii": 4.0,
    "dimplegrinder": 4.0,
    "diamondsaw": 4.0,
    "plasmacleaner": 4.0,
    "grinderpolisher": 4.0,
    "microtome": 4.0,
    "tembiosamples": 12.0,
    "temfibsamplesprep": 12.0,
}

TOOL_MAX_HOURS_ALIASES = {
    "laurellspinner1": "laurellspinner",
    "heidelberg3umlaserwriter": "heidelberg3micronlaserwriter",
    "heidelbergdwl66laserwriter": "heidelbergdw66laserwriter",
    "sussma6duvmaskaligner": "duvma6maskaligner",
    "sussma6maskaligner": "ma6maskaligner",
    "feitalosf200xstem": "feitalostem",
    "zeisssem": "zeisssigmasem",
    "edwardsthermalevaporator1": "edwardsthermalevaporator",
    "angstromhighvacuum": "angstromhighvacuumevaporator",
    "angstrommetalsdepositionsystem": "angstromevovacsystem",
    "ajaorion8dielectricssputteringsystem": "ajadielectricsputter",
    "ajaorion3metalsputteringsystem": "ajametalsputter",
    "uvocsuvozonecleaner": "uvozone",
    "tptwirebonderhb10": "tptwirebonder",
    "cmp": "chemicalmechanicalpolishing",
    "klap17profiler": "klaprofilometer",
    "balteccpd": "criticalpointdryer",
    "micrometricsasap2020hvbetanalyzer": "bet",
    "tosohecosecriuvgpc": "agilentecosecgpc",
    "agilent1260infinitygpc": "agilentecosecgpc",
    "brukerdimensionsfastscanafm": "brukerdimensionfastscanafm",
    "horibaxploramicroraman": "horibamicroraman",
    "phi5500xps": "phixps",
    "rigakusmartlabxrd": "smartlabxrd",
    "rigakuxtalabsynergysscxrd": "synergysscxrd",
    "tembiosamples": "tembiosamples",
    "fibsamplepreparation": "temfibsamplesprep",
}


# -----------------------------
# Parsing helpers
# -----------------------------
def parse_nemo_datetime(s: object) -> Optional[dt.datetime]:
    """Parse strings like '02/16/2026 @ 2:09 PM' into a datetime."""
    if s is None or (isinstance(s, float) and math.isnan(s)):
        return None
    text = str(s).strip()
    for fmt in ("%m/%d/%Y @ %I:%M %p", "%m/%d/%Y@%I:%M %p"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_adjustment_datetime(value: object) -> Optional[dt.datetime]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = dt.datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)
    return parsed.replace(microsecond=0)


def normalize_item(item: object) -> str:
    """Normalize Item names to match tool mapping. Removes ' (Individual)' and ' (Group)' suffixes."""
    if item is None or (isinstance(item, float) and math.isnan(item)):
        return ""
    s = str(item).strip()
    s = re.sub(r"\s*\((Individual|Group)\)\s*$", "", s)
    return s


def normalize_tool_lookup_key(value: object) -> str:
    text = normalize_item(value).lower()
    text = text.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", text)


def normalize_matching_datetime(value: Optional[dt.datetime]) -> Optional[dt.datetime]:
    if value is None:
        return None
    return value.replace(second=0, microsecond=0, tzinfo=None)


def parse_tool_id(value: object) -> Optional[int]:
    if value is None or pd.isna(value):
        return None
    try:
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return None
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
        return int(text)
    except (TypeError, ValueError):
        return None


def parse_hourly_rate_from_rate(rate: object) -> Optional[float]:
    if rate is None or (isinstance(rate, float) and math.isnan(rate)):
        return None
    text = str(rate).strip()
    if not text:
        return None
    match = re.search(r"\$?\s*(\d+(?:\.\d+)?)\s*/\s*hr\b", text, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def compute_adjusted_cost(
    rate: object,
    quantity_minutes: Optional[float],
    original_cost: float,
) -> float:
    if quantity_minutes is None or pd.isna(quantity_minutes):
        return float(original_cost)

    hourly_rate = parse_hourly_rate_from_rate(rate)
    minimum_charge = parse_minimum_charge_from_rate(rate)
    if hourly_rate is None:
        return float(original_cost)

    adjusted_cost = hourly_rate * (float(quantity_minutes) / 60.0)
    if minimum_charge is not None:
        adjusted_cost = max(adjusted_cost, minimum_charge)
    return round(adjusted_cost, 2)


def resolve_billable_user_key(row: pd.Series) -> str:
    username = str(row.get("Username") or "").strip()
    if username:
        return username
    return str(row.get("User") or "").strip()


def resolve_max_billable_hours(row: pd.Series) -> Optional[float]:
    for column in ("Tool ID", "Tool Id", "ToolID", "Tool"):
        if column in row.index:
            tool_id = parse_tool_id(row[column])
            if tool_id is not None and tool_id in TOOL_MAX_HOURS_BY_TOOL_ID:
                return TOOL_MAX_HOURS_BY_TOOL_ID[tool_id]

    tool_key = normalize_tool_lookup_key(row.get("Item"))
    tool_key = TOOL_MAX_HOURS_ALIASES.get(tool_key, tool_key)
    return TOOL_MAX_HOURS_BY_NAME.get(tool_key)


def apply_max_session_charge_caps(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df["Max Billable Hours"] = df.apply(resolve_max_billable_hours, axis=1)
    quantity_hours = df["Quantity"] / 60.0
    capped_mask = (
        ~df["IsConsumable"]
        & df["Max Billable Hours"].notna()
        & df["Quantity"].notna()
        & (quantity_hours > df["Max Billable Hours"])
    )
    if not capped_mask.any():
        return df

    original_cost = df.loc[capped_mask, "Cost"].copy()
    original_quantity = df.loc[capped_mask, "Quantity"].copy()
    capped_quantity = df.loc[capped_mask, "Max Billable Hours"].astype(float) * 60.0

    df.loc[capped_mask, "Original Quantity"] = original_quantity
    df.loc[capped_mask, "Original Cost"] = original_cost
    df.loc[capped_mask, "Quantity"] = capped_quantity

    scaled_cost = original_cost.copy()
    positive_quantity_mask = original_quantity > 0
    scaled_cost.loc[positive_quantity_mask] = (
        original_cost.loc[positive_quantity_mask]
        * capped_quantity.loc[positive_quantity_mask]
        / original_quantity.loc[positive_quantity_mask]
    )
    df.loc[capped_mask, "Cost"] = scaled_cost.round(2)
    return df


def _scale_costs_to_target(costs: pd.Series, target_total: float) -> pd.Series:
    if costs.empty:
        return costs

    positive_mask = costs > 0
    if not positive_mask.any():
        return pd.Series(0.0, index=costs.index, dtype=float)

    positive_costs = costs.loc[positive_mask].astype(float)
    total_cost = float(positive_costs.sum())
    if total_cost <= 0:
        return pd.Series(0.0, index=costs.index, dtype=float)

    target_cents = max(0, int(round(target_total * 100)))
    raw_scaled_cents = positive_costs * target_cents / total_cost
    floor_cents = raw_scaled_cents.apply(math.floor).astype(int)
    remainder_cents = target_cents - int(floor_cents.sum())

    if remainder_cents > 0:
        remainders = (raw_scaled_cents - floor_cents).sort_values(ascending=False)
        for index in remainders.index[:remainder_cents]:
            floor_cents.loc[index] += 1

    scaled = pd.Series(0.0, index=costs.index, dtype=float)
    scaled.loc[positive_mask] = floor_cents / 100.0
    return scaled


def apply_project_charge_caps(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    if "Period" not in df.columns:
        return df

    if "Billable User Key" not in df.columns:
        df["Billable User Key"] = df.apply(resolve_billable_user_key, axis=1)

    capped_row_indexes: list[int] = []

    for (period, user_key, project, application_identifier), index in df.groupby(
        ["Period", "Billable User Key", "Project", "Application identifier"],
        dropna=False,
    ).groups.items():
        project_cap = PROJECT_CHARGE_CAP_BY_APPLICATION.get(str(application_identifier))
        if project_cap is None:
            continue

        group_index = pd.Index(index)
        group_costs = df.loc[group_index, "Cost"].fillna(0.0).astype(float)
        group_total = float(group_costs.sum())
        if group_total <= project_cap:
            continue

        scaled_costs = _scale_costs_to_target(group_costs, project_cap)
        capped_mask = group_costs > scaled_costs
        if not capped_mask.any():
            continue

        capped_indexes = group_index[capped_mask]
        capped_row_indexes.extend(capped_indexes.tolist())

        df.loc[capped_indexes, "Original Project Cost"] = group_costs.loc[capped_mask]
        df.loc[capped_indexes, "Project Cap Applied"] = project_cap
        df.loc[capped_indexes, "Cost"] = scaled_costs.loc[capped_mask]

    if capped_row_indexes:
        df.loc[capped_row_indexes, "Project Cap Reduction"] = (
            df.loc[capped_row_indexes, "Original Project Cost"]
            - df.loc[capped_row_indexes, "Cost"]
        ).round(2)

    return df


def _project_name_from_metadata(project_payload: dict[str, Any]) -> str:
    return str(project_payload.get("name") or "").strip()


def _project_application_identifier_from_metadata(
    project_payload: dict[str, Any],
) -> Optional[str]:
    candidate_keys = (
        "application_identifier",
        "application identifier",
        "account_type",
        "account type",
    )
    for key in candidate_keys:
        value = project_payload.get(key)
        if value:
            return str(value).strip()
    return None


def apply_adjustment_requests(
    df: pd.DataFrame,
    adjustment_requests: list[dict[str, Any]],
    tools_by_id: Dict[int, str],
    projects_by_name: Dict[str, dict[str, Any]],
) -> pd.DataFrame:
    if df.empty or not adjustment_requests or not tools_by_id:
        return df

    project_name_by_id: Dict[int, str] = {}
    project_app_identifier_by_id: Dict[int, str] = {}
    for project_payload in projects_by_name.values():
        project_id = project_payload.get("id")
        if isinstance(project_id, int):
            project_name = _project_name_from_metadata(project_payload)
            if project_name:
                project_name_by_id[project_id] = project_name
            app_identifier = _project_application_identifier_from_metadata(project_payload)
            if app_identifier:
                project_app_identifier_by_id[project_id] = app_identifier

    working = df.copy()
    working["_original_row_index"] = working.index
    working["_tool_lookup_key"] = working["Item"].apply(normalize_tool_lookup_key)
    working["_start_match_dt"] = working["Start time"].apply(parse_nemo_datetime).apply(
        normalize_matching_datetime
    )
    working["_end_match_dt"] = working["End time"].apply(parse_nemo_datetime).apply(
        normalize_matching_datetime
    )

    adjustment_df = pd.DataFrame(adjustment_requests)
    if adjustment_df.empty:
        return df

    approved = adjustment_df[
        adjustment_df.get("status", pd.Series(dtype=float)).eq(1)
        & ~adjustment_df.get("deleted", pd.Series(dtype=bool)).fillna(False)
    ].copy()
    if approved.empty:
        return df

    used_row_indexes: set[int] = set()
    rows_to_drop: set[int] = set()

    for adjustment in approved.to_dict("records"):
        tool_id = parse_tool_id(adjustment.get("item_tool"))
        tool_name = tools_by_id.get(tool_id) if tool_id is not None else None
        if not tool_name:
            continue

        tool_key = normalize_tool_lookup_key(tool_name)
        original_start = normalize_matching_datetime(
            parse_adjustment_datetime(adjustment.get("original_start"))
        )
        original_end = normalize_matching_datetime(
            parse_adjustment_datetime(adjustment.get("original_end"))
        )
        if original_start is None or original_end is None:
            continue

        matches = working[
            (working["_tool_lookup_key"] == tool_key)
            & (working["_start_match_dt"] == original_start)
            & (working["_end_match_dt"] == original_end)
            & ~working["_original_row_index"].isin(used_row_indexes)
        ]
        if matches.empty:
            continue

        if len(matches) > 1:
            original_project_id = parse_tool_id(adjustment.get("original_project"))
            if original_project_id is not None:
                original_project_name = project_name_by_id.get(original_project_id, "")
                if original_project_name:
                    narrowed = matches[matches["Project"] == original_project_name]
                    if not narrowed.empty:
                        matches = narrowed
            if len(matches) > 1:
                matches = matches.sort_values("_original_row_index").iloc[[0]]

        matched_index = int(matches.index[0])
        used_row_indexes.add(int(matches.iloc[0]["_original_row_index"]))

        new_start_dt = parse_adjustment_datetime(adjustment.get("new_start"))
        new_end_dt = parse_adjustment_datetime(adjustment.get("new_end"))
        waive = bool(adjustment.get("waive"))

        if waive or (
            new_start_dt is not None
            and new_end_dt is not None
            and new_start_dt == new_end_dt
        ):
            rows_to_drop.add(matched_index)
            continue

        if new_start_dt is not None and new_end_dt is not None and new_end_dt >= new_start_dt:
            quantity_minutes = (new_end_dt - new_start_dt).total_seconds() / 60.0
            working.at[matched_index, "Start time"] = new_start_dt.strftime("%m/%d/%Y @ %I:%M %p")
            working.at[matched_index, "End time"] = new_end_dt.strftime("%m/%d/%Y @ %I:%M %p")
            working.at[matched_index, "Start_dt"] = new_start_dt
            working.at[matched_index, "Quantity"] = quantity_minutes
            working.at[matched_index, "Cost"] = compute_adjusted_cost(
                working.at[matched_index, "Rate"],
                quantity_minutes,
                float(working.at[matched_index, "Cost"] or 0.0),
            )
            working.at[matched_index, "_start_match_dt"] = normalize_matching_datetime(new_start_dt)
            working.at[matched_index, "_end_match_dt"] = normalize_matching_datetime(new_end_dt)

        new_project_id = parse_tool_id(adjustment.get("new_project"))
        if new_project_id is not None:
            new_project_name = project_name_by_id.get(new_project_id)
            if new_project_name:
                working.at[matched_index, "Project"] = new_project_name
            new_app_identifier = project_app_identifier_by_id.get(new_project_id)
            if new_app_identifier:
                working.at[matched_index, "Application identifier"] = new_app_identifier

    if rows_to_drop:
        working = working.drop(index=list(rows_to_drop))

    working = working.drop(
        columns=["_original_row_index", "_tool_lookup_key", "_start_match_dt", "_end_match_dt"],
        errors="ignore",
    )
    return working.reset_index(drop=True)


def parse_minimum_charge_from_rate(rate: object) -> Optional[float]:
    """
    Extract a minimum charge from the Rate text when present.
    Handles patterns such as "$10/hr ($5 minimum)" or "minimum charge $5".
    """
    if rate is None or (isinstance(rate, float) and math.isnan(rate)):
        return None
    text = str(rate).strip()
    if not text:
        return None

    patterns = (
        r"\$?\s*(\d+(?:\.\d+)?)\s*minimum",
        r"minimum(?:\s+charge)?[^$0-9]*\$?\s*(\d+(?:\.\d+)?)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                return None
    return None


def extract_pi_code_from_project(project: object) -> str:
    """Fallback PI id: take last token from the Project string (often a UNI like ML3745)."""
    if project is None or (isinstance(project, float) and math.isnan(project)):
        return "UNKNOWN_PI"
    tokens = str(project).strip().split()
    return tokens[-1] if tokens else "UNKNOWN_PI"


def period_from_start_dt(start_dt: Optional[dt.datetime]) -> str:
    """Return YYYY-MM for grouping invoices by month."""
    if not start_dt:
        return "Unknown"
    return start_dt.strftime("%Y-%m")


def month_label(period_ym: str) -> str:
    """Convert 'YYYY-MM' to 'MONYYYY' (e.g., '2026-01' -> 'JAN2026')."""
    try:
        y, m = period_ym.split("-")
        month = dt.date(int(y), int(m), 1).strftime("%b").upper()
        return f"{month}{y}"
    except Exception:
        return period_ym


def safe_filename(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\s,\-]+", "", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s or "UNKNOWN_PI"


def make_invoice_number(period_ym: str, seq: int = 1) -> str:
    """
    Stable invoice number format: CNI-YYMM-SEQ
    Example: CNI-2603-001
    """
    period_code = str(period_ym or "").replace("-", "")
    if len(period_code) == 6:
        period_code = period_code[2:]
    return f"CNI-{period_code}-{int(seq):03d}"


# -----------------------------
# NEMO API helpers (optional)
# -----------------------------
@dataclass(frozen=True)
class PIInfo:
    key: str  # stable key used for grouping (email or UNI)
    display_name: str  # 'Last, First' if possible
    email: str = ""


def _requests_required() -> None:
    if requests is None:
        raise RuntimeError(
            "The 'requests' library is required for API access. Install it or run without an API token."
        )


def _nemo_cache_key(kind: str, nemo_base: str, api_token: str) -> tuple[str, str, str]:
    return (kind, nemo_base.rstrip("/"), api_token.strip())


def _load_cached_nemo_metadata(
    kind: str, nemo_base: str, api_token: str
) -> Optional[dict]:
    cache_key = _nemo_cache_key(kind, nemo_base, api_token)
    cached = NEMO_METADATA_CACHE.get(cache_key)
    if not cached:
        return None
    expires_at, payload = cached
    if time.monotonic() >= expires_at:
        NEMO_METADATA_CACHE.pop(cache_key, None)
        return None
    return dict(payload)


def _store_cached_nemo_metadata(
    kind: str, nemo_base: str, api_token: str, payload: dict
) -> None:
    cache_key = _nemo_cache_key(kind, nemo_base, api_token)
    NEMO_METADATA_CACHE[cache_key] = (
        time.monotonic() + NEMO_METADATA_CACHE_TTL_SECONDS,
        dict(payload),
    )


def fetch_all_projects(
    nemo_base: str,
    api_token: str,
    *,
    use_cache: bool = True,
    status_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, dict]:
    """
    Fetch all projects and return a dict {project_name: project_json}.
    Works with DRF pagination ('results' + 'next') and with non-paginated lists.
    """
    _requests_required()
    if use_cache:
        cached_projects = _load_cached_nemo_metadata("projects", nemo_base, api_token)
        if cached_projects is not None:
            if status_callback:
                status_callback("Using cached NEMO project contacts")
            return cached_projects

    url = nemo_base.rstrip("/") + "/api/projects/"
    headers = {"Authorization": f"Token {api_token}"}
    projects: Dict[str, dict] = {}
    page_number = 0

    while url:
        page_number += 1
        if status_callback:
            status_callback(f"Fetching NEMO project contacts (page {page_number})")
        r = requests.get(url, headers=headers, timeout=60)
        r.raise_for_status()
        payload = r.json()

        if isinstance(payload, list):
            results = payload
            url = None
        else:
            results = payload.get("results", [])
            url = payload.get("next")

        for p in results:
            name = p.get("name")
            if name:
                projects[name] = p

    _store_cached_nemo_metadata("projects", nemo_base, api_token, projects)
    return projects


def fetch_all_tools(
    nemo_base: str,
    api_token: str,
    *,
    use_cache: bool = True,
    status_callback: Optional[Callable[[str], None]] = None,
) -> Dict[int, str]:
    """Fetch all tools and return a dict {tool_id: tool_name}."""
    _requests_required()
    if use_cache:
        cached_tools = _load_cached_nemo_metadata("tools", nemo_base, api_token)
        if cached_tools is not None:
            if status_callback:
                status_callback("Using cached NEMO tools metadata")
            return {
                int(key): str(value) for key, value in cached_tools.items() if str(key).isdigit()
            }

    url = nemo_base.rstrip("/") + "/api/tools/"
    headers = {"Authorization": f"Token {api_token}"}
    tools: Dict[int, str] = {}
    page_number = 0

    while url:
        page_number += 1
        if status_callback:
            status_callback(f"Fetching NEMO tools metadata (page {page_number})")
        r = requests.get(url, headers=headers, timeout=60)
        r.raise_for_status()
        payload = r.json()

        if isinstance(payload, list):
            results = payload
            url = None
        else:
            results = payload.get("results", [])
            url = payload.get("next")

        for tool in results:
            tool_id = tool.get("id")
            tool_name = tool.get("name")
            if isinstance(tool_id, int) and tool_name:
                tools[tool_id] = str(tool_name)

    _store_cached_nemo_metadata("tools", nemo_base, api_token, {str(k): v for k, v in tools.items()})
    return tools


def fetch_all_adjustment_requests(
    nemo_base: str,
    api_token: str,
    *,
    use_cache: bool = True,
    status_callback: Optional[Callable[[str], None]] = None,
) -> list[dict[str, Any]]:
    """Fetch all adjustment requests."""
    _requests_required()
    if use_cache:
        cached_adjustments = _load_cached_nemo_metadata(
            "adjustment_requests", nemo_base, api_token
        )
        if cached_adjustments is not None:
            if status_callback:
                status_callback("Using cached NEMO adjustment requests")
            records = cached_adjustments.get("results", [])
            if isinstance(records, list):
                return [record for record in records if isinstance(record, dict)]

    url = nemo_base.rstrip("/") + "/api/adjustment_requests/"
    headers = {"Authorization": f"Token {api_token}"}
    adjustments: list[dict[str, Any]] = []
    page_number = 0

    while url:
        page_number += 1
        if status_callback:
            status_callback(f"Fetching NEMO adjustment requests (page {page_number})")
        r = requests.get(url, headers=headers, timeout=60)
        r.raise_for_status()
        payload = r.json()

        if isinstance(payload, list):
            results = payload
            url = None
        else:
            results = payload.get("results", [])
            url = payload.get("next")

        for adjustment in results:
            if isinstance(adjustment, dict):
                adjustments.append(adjustment)

    _store_cached_nemo_metadata(
        "adjustment_requests",
        nemo_base,
        api_token,
        {"results": adjustments},
    )
    return adjustments


def lab_for_consumable_category(category: object) -> Optional[str]:
    """
    Map NEMO consumable categories to invoice labs.
    Billing rule:
    - category 4 -> SMCL
    - categories 1, 2, 3 -> Cleanroom
    """
    try:
        category_id = int(category)
    except (TypeError, ValueError):
        return None

    if category_id == 4:
        return "SMCL"
    if category_id in (1, 2, 3):
        return "Cleanroom"
    return None


def fetch_all_consumables(
    nemo_base: str,
    api_token: str,
    *,
    use_cache: bool = True,
    status_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, str]:
    """
    Fetch all consumables and return a dict {normalized_consumable_name: lab_name}.
    """
    _requests_required()
    if use_cache:
        cached_consumables = _load_cached_nemo_metadata(
            "consumables", nemo_base, api_token
        )
        if cached_consumables is not None:
            if status_callback:
                status_callback("Using cached NEMO consumable metadata")
            return {str(key): str(value) for key, value in cached_consumables.items()}

    url = nemo_base.rstrip("/") + "/api/consumables/"
    headers = {"Authorization": f"Token {api_token}"}
    consumable_labs: Dict[str, str] = {}
    page_number = 0

    while url:
        page_number += 1
        if status_callback:
            status_callback(f"Fetching NEMO consumable data (page {page_number})")
        r = requests.get(url, headers=headers, timeout=60)
        r.raise_for_status()
        payload = r.json()

        if isinstance(payload, list):
            results = payload
            url = None
        else:
            results = payload.get("results", [])
            url = payload.get("next")

        for consumable in results:
            name = normalize_item(consumable.get("name"))
            lab = lab_for_consumable_category(consumable.get("category"))
            if name and lab:
                consumable_labs[name] = lab

    _store_cached_nemo_metadata("consumables", nemo_base, api_token, consumable_labs)
    return consumable_labs


def guess_last_first(name: str) -> str:
    """Turn 'First Last' -> 'Last, First'. If already contains a comma, keep as-is."""
    name = (name or "").strip()
    if not name:
        return ""
    if "," in name:
        return name
    parts = name.split()
    if len(parts) == 1:
        return parts[0]
    return f"{parts[-1]}, {' '.join(parts[:-1])}"


def resolve_pi_for_project(project_name: str, project_map: Dict[str, dict]) -> PIInfo:
    """Use project JSON to choose PI grouping key and display name."""
    p = project_map.get(project_name)
    if not p:
        code = extract_pi_code_from_project(project_name)
        return PIInfo(key=code, display_name=code, email="")

    email = (p.get("contact_email") or "").strip().lower()
    cname = guess_last_first(p.get("contact_name") or "")
    if email:
        return PIInfo(key=email, display_name=cname or email, email=email)

    code = extract_pi_code_from_project(project_name)
    return PIInfo(key=code, display_name=cname or code, email="")


def internal_facility_fee_for_group(df_group: pd.DataFrame) -> float:
    """
    Pick the invoice-level internal fee based on the application identifiers present.
    If multiple types are present, use the highest configured fee.
    """
    app_ids = set(
        df_group.get("Application identifier", pd.Series(dtype=str))
        .dropna()
        .astype(str)
    )
    if not app_ids:
        return 0.0
    return float(
        max(INTERNAL_FACILITY_FEE_BY_APPLICATION.get(app_id, 0.0) for app_id in app_ids)
    )


def invoice_group_has_cdg(df_group: pd.DataFrame) -> bool:
    """Return True when the invoice includes any CDG account type charges."""
    app_ids = df_group.get("Application identifier", pd.Series(dtype=str))
    if app_ids.empty:
        return False
    return app_ids.fillna("").astype(str).str.upper().eq("CDG").any()


def select_access_fee_project(df_group: pd.DataFrame) -> Optional[pd.Series]:
    """
    Choose which project should carry the access fee.
    Rule:
    - Use the project with the highest total charges for the billing month
    """
    if df_group.empty:
        return None

    proj_usage = (
        df_group.groupby(["Project", "Application identifier"], dropna=False)["Cost"]
        .sum()
        .reset_index(name="Usage Charges")
    )
    if proj_usage.empty:
        return None

    proj_usage = proj_usage.sort_values(
        ["Usage Charges", "Project"],
        ascending=[False, True],
    ).reset_index(drop=True)
    return proj_usage.iloc[0]


def _is_consumable_type(value: object) -> bool:
    return "consum" in str(value or "").strip().lower()


# -----------------------------
# Excel writer
# -----------------------------
_THIN = Side(style="thin", color="000000")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue
_SECTION_FILL = PatternFill("solid", fgColor="FCE4D6")  # light orange
_BOLD = Font(bold=True)
_TITLE = Font(bold=True, size=14)
_TITLE_LARGE = Font(bold=True, size=18)


def autosize_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            if isinstance(v, (dt.date, dt.datetime)):
                s = v.strftime("%Y-%m-%d %H:%M")
            else:
                s = str(v)
            max_len = max(max_len, len(s))
        width = min(max_width, max(min_width, max_len + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def write_table(
    ws,
    start_row: int,
    start_col: int,
    df_table: pd.DataFrame,
    currency_cols: Iterable[str] = (),
) -> int:
    currency_cols = set(currency_cols)

    # headers
    for j, colname in enumerate(df_table.columns, start=start_col):
        cell = ws.cell(start_row, j, value=colname)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.border = _BORDER_THIN
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # rows
    for i, (_, row) in enumerate(df_table.iterrows(), start=1):
        for j, colname in enumerate(df_table.columns, start=start_col):
            val = row[colname]
            cell = ws.cell(start_row + i, j, value=val)
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if colname in currency_cols:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    return start_row + len(df_table) + 1


def create_invoice_workbook(
    df_group: pd.DataFrame,
    pi_display_name: str,
    period_ym: str,
    invoice_number: str,
    pi_email: str = "",
) -> Workbook:
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.calcOnSave = True
    wb.calculation.calcCompleted = False
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "Invoice"

    ml = month_label(period_ym)
    internal_fee = internal_facility_fee_for_group(df_group)
    show_subsidy = invoice_group_has_cdg(df_group)
    detail_end_col = 9 if show_subsidy else 8

    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "Columbia Nano Initiative"
    ws["A1"].font = _TITLE_LARGE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = "Facility Usage Invoice"
    ws["A2"].font = _TITLE
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "PI"
    ws["B4"] = pi_display_name
    ws["D4"] = "Billing Month"
    ws["E4"] = ml
    ws["G4"] = "Generated"
    ws["H4"] = invoice_generated_at().replace(tzinfo=None)
    ws["J4"] = "Invoice #"
    ws["K4"] = invoice_number
    ws["A5"] = "Email"
    ws["B5"] = pi_email or "N/A"

    for cell in ("A4", "D4", "G4", "J4", "A5"):
        ws[cell].font = _BOLD
    ws["B4"].font = _TITLE
    ws["E4"].font = _TITLE
    ws["H4"].number_format = "yyyy-mm-dd hh:mm"
    ws["K4"].font = _TITLE

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    # Summary totals by lab
    summary_start = 6
    ws.cell(summary_start, 1, value="Lab").font = _BOLD
    ws.cell(summary_start, 2, value="Total Cost").font = _BOLD
    for c in (1, 2):
        ws.cell(summary_start, c).fill = _SECTION_FILL
        ws.cell(summary_start, c).border = _BORDER_THIN

    lab_tot = df_group.groupby("Lab")["Cost"].sum().sort_index()
    summary_lab_rows: Dict[str, int] = {}

    r = summary_start + 1
    for lab, cost in lab_tot.items():
        ws.cell(r, 1, value=lab).border = _BORDER_THIN
        summary_lab_rows[str(lab)] = r
        ccell = ws.cell(r, 2, value=float(cost))
        ccell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ccell.border = _BORDER_THIN
        r += 1

    # Add access fee as a separate line item (once per invoice)
    ws.cell(r, 1, value="Access fee").border = _BORDER_THIN
    summary_access_fee_row = r
    fee_cell = ws.cell(r, 2, value=internal_fee)
    fee_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    fee_cell.border = _BORDER_THIN
    r += 1

    ws.cell(r, 1, value="TOTAL").font = _BOLD
    ws.cell(r, 1).border = _BORDER_THIN
    summary_total_row = r
    total_cell = ws.cell(r, 2, value=float(lab_tot.sum() + internal_fee))
    total_cell.font = _BOLD
    total_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    total_cell.border = _BORDER_THIN

    # Details by lab
    current_row = r + 2
    lab_subtotal_cells: Dict[str, str] = {}
    detail_sections: List[dict[str, object]] = []

    for lab in DESIRED_LAB_ORDER:
        df_lab = df_group[df_group["Lab"] == lab].copy()
        if df_lab.empty:
            continue

        title = ws.cell(current_row, 1, value=lab)
        title.font = Font(bold=True, size=12)
        title.fill = _SECTION_FILL
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=detail_end_col,
        )
        current_row += 1

        df_lab = df_lab.sort_values(["Start_dt", "User", "Item_norm", "Project"])

        detail_columns = [
            "Start_dt",
            "User",
            "Item_norm",
            "Type",
            "Project",
            "Quantity",
            "Rate",
        ]
        if show_subsidy:
            detail_columns.append("Subsidy")
        detail_columns.append("Cost")

        out = df_lab[detail_columns].copy()
        out = out.rename(columns={"Start_dt": "Date", "Item_norm": "Description"})

        currency_cols = ["Cost"]
        if show_subsidy:
            currency_cols.insert(0, "Subsidy")
        detail_header_row = current_row
        detail_data_start_row = detail_header_row + 1
        detail_data_end_row = detail_header_row + len(out)
        current_row = write_table(
            ws, current_row, 1, out, currency_cols=currency_cols
        )
        cost_col_letter = get_column_letter(detail_end_col)
        project_col_letter = get_column_letter(5)
        detail_sections.append(
            {
                "lab": lab,
                "project_range": (
                    f"${project_col_letter}${detail_data_start_row}:"
                    f"${project_col_letter}${detail_data_end_row}"
                ),
                "cost_range": (
                    f"${cost_col_letter}${detail_data_start_row}:"
                    f"${cost_col_letter}${detail_data_end_row}"
                ),
            }
        )

        subtotal_label_col = detail_end_col - 1
        ws.cell(current_row, subtotal_label_col, value="Subtotal").font = _BOLD
        sub = ws.cell(
            current_row,
            detail_end_col,
            value=f"=SUM({cost_col_letter}{detail_data_start_row}:{cost_col_letter}{detail_data_end_row})",
        )
        sub.font = _BOLD
        sub.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        lab_subtotal_cells[lab] = sub.coordinate
        current_row += 2

    for lab, summary_row in summary_lab_rows.items():
        subtotal_cell = lab_subtotal_cells.get(lab)
        if subtotal_cell:
            ws.cell(summary_row, 2, value=f"={subtotal_cell}")

    ws.cell(
        summary_total_row,
        2,
        value=f"=SUM(B{summary_start + 1}:B{summary_access_fee_row})",
    )

    # Access fee + project fees summary (once per invoice)
    if not df_group.empty:
        access_fee_project = select_access_fee_project(df_group)
        access_fee_detail_cell = ""
        access_fee_project_ref = ""
        access_fee_project_type_ref = ""
        if access_fee_project is not None:
            title = ws.cell(current_row, 1, value="Access fee")
            title.font = Font(bold=True, size=12)
            title.fill = _SECTION_FILL
            ws.merge_cells(
                start_row=current_row, start_column=1, end_row=current_row, end_column=4
            )
            current_row += 1

            access_fee_df = pd.DataFrame(
                [
                    {
                        "Project": access_fee_project["Project"],
                        "Project Type": access_fee_project["Application identifier"],
                        "Access Fee": internal_fee,
                    }
                ]
            )
            current_row = write_table(
                ws,
                current_row,
                1,
                access_fee_df,
                currency_cols=["Access Fee"],
            )
            access_fee_detail_cell = f"C{current_row - 1}"
            access_fee_project_ref = f"$A${current_row - 1}"
            access_fee_project_type_ref = f"$B${current_row - 1}"
            ws.cell(summary_access_fee_row, 2, value=f"={access_fee_detail_cell}")
            ws.row_dimensions[current_row].height = 12
            current_row += 1

        title = ws.cell(current_row, 1, value="Project fees summary")
        title.font = Font(bold=True, size=12)
        title.fill = _SECTION_FILL
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=9
        )
        current_row += 1

        proj_usage = (
            df_group[~df_group["IsConsumable"]]
            .pivot_table(
                index=["Project", "Application identifier"],
                columns="Lab",
                values="Cost",
                aggfunc="sum",
                fill_value=0.0,
            )
            .reset_index()
        )
        proj_consumables = (
            df_group[df_group["IsConsumable"]]
            .groupby(["Project", "Application identifier"], dropna=False)["Cost"]
            .sum()
            .reset_index(name="Consumable")
        )
        proj = proj_usage.merge(
            proj_consumables, on=["Project", "Application identifier"], how="outer"
        ).fillna(0.0)

        for lab in ("Cleanroom", "SMCL", "Electron Microscopy Lab"):
            if lab not in proj.columns:
                proj[lab] = 0.0
        if "Consumable" not in proj.columns:
            proj["Consumable"] = 0.0

        proj["Access Fee"] = 0.0
        if access_fee_project is not None:
            fee_mask = (proj["Project"] == access_fee_project["Project"]) & (
                proj["Application identifier"]
                == access_fee_project["Application identifier"]
            )
            proj.loc[fee_mask, "Access Fee"] = internal_fee
        proj["Project Total"] = (
            proj[["Cleanroom", "SMCL", "Electron Microscopy Lab", "Consumable"]].sum(
                axis=1
            )
            + proj["Access Fee"]
        )
        proj = proj.rename(columns={"Application identifier": "Project Type"})

        proj_cols = (
            ["Project", "Project Type"]
            + list(DESIRED_LAB_ORDER)
            + ["Access Fee", "Project Total"]
        )
        proj = proj[proj_cols].sort_values(
            ["Project Total", "Project"], ascending=[False, True]
        )

        project_summary_header_row = current_row
        project_summary_first_row = project_summary_header_row + 1
        current_row = write_table(
            ws,
            current_row,
            1,
            proj,
            currency_cols=list(DESIRED_LAB_ORDER) + ["Access Fee", "Project Total"],
        )
        project_summary_last_row = current_row - 1

        last_col = len(proj_cols)
        lab_col_by_name = {
            lab: proj_cols.index(lab) + 1 for lab in DESIRED_LAB_ORDER
        }
        access_fee_col = proj_cols.index("Access Fee") + 1
        project_total_col = proj_cols.index("Project Total") + 1

        for row_idx in range(project_summary_first_row, project_summary_last_row + 1):
            project_cell = f"$A{row_idx}"
            for lab in DESIRED_LAB_ORDER:
                terms = [
                    (
                        f'SUMIFS({section["cost_range"]},'
                        f'{section["project_range"]},{project_cell})'
                    )
                    for section in detail_sections
                    if section["lab"] == lab
                ]
                formula = "=" + "+".join(terms) if terms else "=0"
                ws.cell(row_idx, lab_col_by_name[lab], value=formula)

            access_fee_cell = ws.cell(row_idx, access_fee_col)
            if access_fee_detail_cell:
                access_fee_cell.value = (
                    f"=IF(AND($A{row_idx}={access_fee_project_ref},"
                    f"$B{row_idx}={access_fee_project_type_ref}),"
                    f"{access_fee_detail_cell},0)"
                )
            else:
                access_fee_cell.value = "=0"
            access_fee_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            first_lab_col = get_column_letter(lab_col_by_name[DESIRED_LAB_ORDER[0]])
            last_lab_col = get_column_letter(lab_col_by_name[DESIRED_LAB_ORDER[-1]])
            access_fee_col_letter = get_column_letter(access_fee_col)
            project_total_cell = ws.cell(row_idx, project_total_col)
            project_total_cell.value = (
                f"=SUM({first_lab_col}{row_idx}:{access_fee_col_letter}{row_idx})"
            )
            project_total_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        ws.cell(current_row, 1, value="Usage charges total").font = _BOLD
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=last_col - 1,
        )
        first_lab_col = get_column_letter(lab_col_by_name[DESIRED_LAB_ORDER[0]])
        last_lab_col = get_column_letter(lab_col_by_name[DESIRED_LAB_ORDER[-1]])
        tot_cell = ws.cell(
            current_row,
            last_col,
            value=(
                f"=SUM({first_lab_col}{project_summary_first_row}:"
                f"{last_lab_col}{project_summary_last_row})"
            ),
        )
        tot_cell.font = _BOLD
        tot_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        usage_total_cell = tot_cell.coordinate
        current_row += 1

        ws.cell(current_row, 1, value="Access fee").border = _BORDER_THIN
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=last_col - 1,
        )
        access_fee_col_letter = get_column_letter(access_fee_col)
        fee_cell2 = ws.cell(
            current_row,
            last_col,
            value=(
                f"=SUM({access_fee_col_letter}{project_summary_first_row}:"
                f"{access_fee_col_letter}{project_summary_last_row})"
            ),
        )
        fee_cell2.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        bottom_access_fee_cell = fee_cell2.coordinate
        current_row += 1

        ws.cell(current_row, 1, value="Invoice total").font = _BOLD
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=last_col - 1,
        )
        inv_cell = ws.cell(
            current_row,
            last_col,
            value=f"=SUM({usage_total_cell}:{bottom_access_fee_cell})",
        )
        inv_cell.font = _BOLD
        inv_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        current_row += 1

    ws.freeze_panes = "A8"
    autosize_columns(ws)
    return wb


# -----------------------------
# PDF writer
# -----------------------------
def _pdf_available() -> bool:
    return SimpleDocTemplate is not None


def _fmt_money(x: float) -> str:
    try:
        return f"${float(x):,.2f}"
    except Exception:
        return "$0.00"


def _make_logo_flowable(logo_path: str, max_w: float, max_h: float):
    """
    Create a reportlab Image flowable scaled to fit within max_w x max_h (points).
    Returns None if logo_path doesn't exist or reportlab image utilities are missing.
    """
    if not _pdf_available() or not logo_path:
        return None
    if not os.path.exists(logo_path):
        return None
    try:
        reader = ImageReader(logo_path)
        iw, ih = reader.getSize()
        if iw <= 0 or ih <= 0:
            return None
        scale = min(max_w / iw, max_h / ih)
        scale = min(scale, 1.0)  # never upscale
        return Image(logo_path, width=iw * scale, height=ih * scale)
    except Exception:
        return None


def _on_page(canvas, doc):
    """Footer with page number."""
    try:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColorRGB(0.2, 0.2, 0.2)
        canvas.drawRightString(
            doc.pagesize[0] - doc.rightMargin,
            0.35 * inch,
            f"Page {canvas.getPageNumber()}",
        )
        canvas.restoreState()
    except Exception:
        pass


def create_invoice_pdf(
    df_group: pd.DataFrame,
    pi_display_name: str,
    pi_email: str,
    period_ym: str,
    invoice_number: str,
    pdf_path: str,
    logo_path: Optional[str] = None,
    internal_fee_override: Optional[float] = None,
) -> None:
    """
    Create a PDF invoice that mirrors the XLSX content at a high level:
    - Header with optional logo
    - Summary totals
    - Tables per lab
    - Project fees summary
    """
    if not _pdf_available():
        raise RuntimeError(
            "PDF generation requested, but 'reportlab' is not installed."
        )

    # PDF setup
    pagesize = landscape(letter)
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=pagesize,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.6 * inch,
        title="NEMO Invoice",
        author="NEMO Invoice Generator",
    )

    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.fontSize = 14
    styleN.leading = 14

    styleSmall = ParagraphStyle(
        "small",
        parent=styleN,
        fontSize=8,
        leading=10,
    )
    styleSmallBold = ParagraphStyle(
        "smallbold",
        parent=styleSmall,
        fontName="Helvetica-Bold",
    )
    styleTitle = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontSize=16,
        leading=18,
        spaceAfter=6,
    )
    styleH = ParagraphStyle(
        "heading",
        parent=styles["Heading2"],
        fontSize=11,
        leading=13,
        spaceBefore=10,
        spaceAfter=4,
    )

    def P(text: str, st=styleN):
        # Escape any XML-ish chars so project names like "A & B" don't break Paragraph
        return Paragraph(_xml_escape(text or ""), st)

    story: List[object] = []

    # Header (logo + title/info)
    ml = month_label(period_ym)
    internal_fee = (
        float(internal_fee_override)
        if internal_fee_override is not None
        else internal_facility_fee_for_group(df_group)
    )
    show_subsidy = invoice_group_has_cdg(df_group)

    logo = _make_logo_flowable(
        logo_path or "",
        max_w=2.2 * inch,
        max_h=0.9 * inch,
    )
    if logo is None:
        # If no logo, show a text placeholder so the layout stays consistent
        logo_cell: object = P(
            "Columbia University",
            ParagraphStyle("logotxt", parent=styleN, fontSize=14, leading=16),
        )
    else:
        logo.hAlign = "RIGHT"
        logo_cell = logo

    styleTitleCenter = ParagraphStyle(
        "TitleCenter",
        parent=styleTitle,
        alignment=TA_CENTER,
    )

    styleNLeft = ParagraphStyle(
        "NLeft",
        parent=styleN,
        alignment=TA_LEFT,
    )

    # styles
    styleTitleCenter = ParagraphStyle(
        "TitleCenter", parent=styleTitle, alignment=TA_CENTER
    )
    styleNLeft = ParagraphStyle("NLeft", parent=styleN, alignment=TA_LEFT)
    styleNLeftBold = ParagraphStyle(
        "NLeftBold",
        parent=styleNLeft,
        fontName="Helvetica-Bold",  # or Times-Bold
        leading=11,
        spaceBefore=0,
        spaceAfter=0,
    )

    app_ids = set(df_group["Application identifier"].dropna().astype(str))
    show_payment_notice = bool(app_ids.intersection({"Industry", "External Academia"}))
    if show_payment_notice:
        styleRightBold = ParagraphStyle(
            "RightBoldNotice",
            parent=styleSmall,
            fontName="Helvetica-Bold",
            fontSize=12,
            alignment=TA_LEFT,
            leading=12,
        )
        styleRightItalic = ParagraphStyle(
            "RightItalicNotice",
            parent=styleSmall,
            fontName="Helvetica-Oblique",
            fontSize=12,
            alignment=TA_LEFT,
            leading=12,
        )
        payment_bold = Paragraph(
            "<br/>"
            "Please Mail Checks To:<br/>"
            "Columbia Nano Initiative<br/>"
            "530 W 120th Street, RM 1001<br/>"
            "Mail Code 8903 - CEPSR Building<br/>"
            "New York, NY 10027<br/>"
            "Email: cnibilling@columbia.edu",
            styleRightBold,
        )
        payment_italic = Paragraph(
            "Checks Only! Make payable to Columbia University.<br/>"
            "Payment due within 30 days of receipt.",
            styleRightItalic,
        )
        logo_cell = [
            logo_cell,
            Spacer(1, 8),
            payment_bold,
            Spacer(1, 8),
            payment_italic,
        ]

    header_tbl = Table(
        [
            [
                Spacer(1, 1),
                Paragraph(
                    '<font size="18"><b>Columbia Nano Initiative</b></font><br/>'
                    '<font size="15"><b>Facility Usage Invoice</b></font>',
                    styleTitleCenter,
                ),
                logo_cell,
            ],
            [P(f"PI: {pi_display_name}", styleNLeftBold), "", ""],
            [P(f"Email: {pi_email or 'N/A'}", styleNLeftBold), "", ""],
            [P(f"Billing Month: {ml}", styleNLeftBold), "", ""],
            [P(f"Invoice #: {invoice_number}", styleNLeftBold), "", ""],
            [
                P(
                    f"Generated: {invoice_generated_at().strftime('%Y-%m-%d %H:%M')} ET",
                    styleNLeftBold,
                ),
                "",
                "",
            ],
        ],
        colWidths=[
            (doc.width - 2.6 * inch) / 2,
            (doc.width - 2.6 * inch) / 2,
            2.6 * inch,
        ],
        style=TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (1, -1), 0),
                ("BOTTOMPADDING", (2, 0), (2, -1), 2),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),  # title in center column
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),  # logo right
                ("SPAN", (0, 1), (1, 1)),  # detail rows span left+middle
                ("SPAN", (0, 2), (1, 2)),
                ("SPAN", (0, 3), (1, 3)),
                ("SPAN", (0, 4), (1, 4)),
                ("SPAN", (0, 5), (1, 5)),
                ("ALIGN", (0, 1), (1, 5), "LEFT"),  # details left-aligned
                ("SPAN", (2, 0), (2, 5)),  # logo cell spans all rows
            ]
        ),
    )

    story.append(header_tbl)
    story.append(Spacer(1, 8))

    # Summary table
    lab_tot = df_group.groupby("Lab")["Cost"].sum().sort_index()
    usage_total = float(lab_tot.sum())
    invoice_total = usage_total + internal_fee

    summary_data = [["Lab", "Total Cost"]]
    for lab, cost in lab_tot.items():
        summary_data.append([lab, _fmt_money(cost)])
    summary_data.append(["Access fee", _fmt_money(internal_fee)])
    summary_data.append(["TOTAL", _fmt_money(invoice_total)])

    summary_tbl = Table(
        summary_data,
        colWidths=[2.6 * inch, 1.4 * inch],
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FCE4D6")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        ),
        hAlign="LEFT",
    )
    story.append(summary_tbl)
    story.append(Spacer(1, 10))

    # Details by lab
    detail_col_names = ["Date", "User", "Description", "Type", "Project", "Qty", "Rate"]
    if show_subsidy:
        detail_col_names.append("Subsidy")
        # Column widths as fractions of the available page width (doc.width).
        # This prevents tables from being clipped off the right edge.
        detail_fracs = [0.10, 0.10, 0.20, 0.08, 0.19, 0.06, 0.10, 0.08, 0.09]
    else:
        detail_fracs = [0.10, 0.10, 0.22, 0.08, 0.21, 0.06, 0.11]
    detail_col_names.append("Cost")
    col_widths = [doc.width * f for f in detail_fracs]

    for lab in DESIRED_LAB_ORDER:
        df_lab = df_group[df_group["Lab"] == lab].copy()
        if df_lab.empty:
            continue

        story.append(P(lab, styleH))

        df_lab = df_lab.sort_values(["Start_dt", "User", "Item_norm", "Project"])

        rows = [detail_col_names]
        for r in df_lab.itertuples(index=False):
            start_dt = getattr(r, "Start_dt", None)
            date_str = ""
            if isinstance(start_dt, dt.datetime):
                date_str = start_dt.strftime("%Y-%m-%d %H:%M")
            user = str(getattr(r, "User", "") or "")
            desc = str(getattr(r, "Item_norm", "") or "")
            typ = str(getattr(r, "Type", "") or "")
            proj = str(getattr(r, "Project", "") or "")
            qty = getattr(r, "Quantity", "")
            rate = str(getattr(r, "Rate", "") or "")
            subsidy = float(getattr(r, "Subsidy", 0.0) or 0.0)
            cost = float(getattr(r, "Cost", 0.0) or 0.0)

            # Wrap long fields with Paragraphs
            rows.append(
                [
                    P(date_str, styleSmall),
                    P(user, styleSmall),
                    P(desc, styleSmall),
                    P(typ, styleSmall),
                    P(proj, styleSmall),
                    P("" if pd.isna(qty) else f"{qty:g}", styleSmall),
                    P(rate, styleSmall),
                ]
            )
            if show_subsidy:
                rows[-1].append(P(_fmt_money(subsidy), styleSmall))
            rows[-1].append(P(_fmt_money(cost), styleSmall))

        tbl = Table(
            rows,
            colWidths=col_widths,
            repeatRows=1,
            hAlign="LEFT",
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (-2, 1), (-1, -1), "RIGHT"),
                    ("ALIGN", (5, 1), (5, -1), "RIGHT"),
                ]
            ),
        )
        story.append(tbl)

        subtotal = float(df_lab["Cost"].sum())
        subtotal_row = [""] * len(detail_col_names)
        subtotal_row[-2] = "Subtotal"
        subtotal_row[-1] = _fmt_money(subtotal)
        sub_tbl = Table(
            [subtotal_row],
            colWidths=col_widths,
            style=TableStyle(
                [
                    ("FONTNAME", (len(detail_col_names) - 2, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (len(detail_col_names) - 1, 0), (len(detail_col_names) - 1, 0), "RIGHT"),
                    (
                        "LINEABOVE",
                        (len(detail_col_names) - 2, 0),
                        (len(detail_col_names) - 1, 0),
                        0.5,
                        colors.black,
                    ),
                ]
            ),
            hAlign="LEFT",
        )
        story.append(sub_tbl)
        story.append(Spacer(1, 8))

    access_fee_project = select_access_fee_project(df_group)
    if access_fee_project is not None:
        story.append(P("Access fee", styleH))
        access_fee_rows = [
            [
                Paragraph("Project", styleSmallBold),
                Paragraph("Project Type", styleSmallBold),
                Paragraph("Access Fee", styleSmallBold),
            ],
            [
                P(str(access_fee_project["Project"] or ""), styleSmall),
                P(str(access_fee_project["Application identifier"] or ""), styleSmall),
                P(_fmt_money(internal_fee), styleSmall),
            ],
        ]
        access_fee_tbl = Table(
            access_fee_rows,
            colWidths=[doc.width * 0.64, doc.width * 0.16, doc.width * 0.20],
            hAlign="LEFT",
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ]
            ),
        )
        story.append(access_fee_tbl)
        story.append(Spacer(1, 8))

    # Project fees summary
    story.append(P("Project fees summary", styleH))

    proj_usage = (
        df_group[~df_group["IsConsumable"]]
        .pivot_table(
            index=["Project", "Application identifier"],
            columns="Lab",
            values="Cost",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    if "Project" not in proj_usage.columns:
        proj_usage = pd.DataFrame(columns=["Project", "Application identifier"])
    proj_consumables = (
        df_group[df_group["IsConsumable"]]
        .groupby(["Project", "Application identifier"], dropna=False)["Cost"]
        .sum()
        .reset_index(name="Consumable")
    )
    if "Project" not in proj_consumables.columns:
        proj_consumables = pd.DataFrame(
            columns=["Project", "Application identifier", "Consumable"]
        )
    proj = proj_usage.merge(
        proj_consumables, on=["Project", "Application identifier"], how="outer"
    ).fillna(0.0)

    for lab in ("Cleanroom", "SMCL", "Electron Microscopy Lab"):
        if lab not in proj.columns:
            proj[lab] = 0.0
    if "Consumable" not in proj.columns:
        proj["Consumable"] = 0.0
    proj["Access Fee"] = 0.0
    if access_fee_project is not None:
        fee_mask = (proj["Project"] == access_fee_project["Project"]) & (
            proj["Application identifier"]
            == access_fee_project["Application identifier"]
        )
        proj.loc[fee_mask, "Access Fee"] = internal_fee
    proj["Project Total"] = (
        proj[["Cleanroom", "SMCL", "Electron Microscopy Lab", "Consumable"]].sum(axis=1)
        + proj["Access Fee"]
    )

    proj_cols = (
        ["Project", "Application identifier"]
        + list(DESIRED_LAB_ORDER)
        + ["Access Fee", "Project Total"]
    )
    proj = proj[proj_cols].sort_values(
        ["Project Total", "Project"], ascending=[False, True]
    )

    # Use Paragraphs for the header so long lab names wrap instead of overlapping.
    proj_header = [
        Paragraph("Project", styleSmallBold),
        Paragraph("Project Type", styleSmallBold),
    ]
    for lab in DESIRED_LAB_ORDER:
        proj_header.append(Paragraph(lab, styleSmallBold))
    proj_header.append(Paragraph("Access Fee", styleSmallBold))
    proj_header.append(Paragraph("Project Total", styleSmallBold))
    proj_rows = [proj_header]
    for r in proj.itertuples(index=False):
        # r = Project, Application identifier, labs..., Project Total
        vals = list(r)
        project_name = str(vals[0] or "")
        app_id = str(vals[1] or "")
        lab_costs = vals[2:-2]
        access_fee = vals[-2]
        ptotal = vals[-1]

        proj_rows.append(
            [
                P(project_name, styleSmall),
                P(app_id, styleSmall),
                *[P(_fmt_money(v), styleSmall) for v in lab_costs],
                P(_fmt_money(access_fee), styleSmall),
                P(_fmt_money(ptotal), styleSmall),
            ]
        )

    # Column widths for project summary (landscape)
    # Column widths for project summary (landscape) as fractions of doc.width.
    # Columns: Project, Account Type, Cleanroom, SMCL, Electron Microscopy Lab, Consumable, Access Fee, Project Total
    proj_fracs = [0.36, 0.09, 0.09, 0.07, 0.11, 0.09, 0.09, 0.10]
    proj_col_widths = [doc.width * f for f in proj_fracs]
    # If labs order changes, widths might not match; guard
    if len(proj_col_widths) != len(proj_header):
        # fallback: distribute evenly
        proj_col_widths = [doc.width / len(proj_header)] * len(proj_header)

    proj_tbl = Table(
        proj_rows,
        colWidths=proj_col_widths,
        repeatRows=1,
        hAlign="LEFT",
        style=TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ]
        ),
    )
    story.append(proj_tbl)
    story.append(Spacer(1, 8))

    totals_tbl = Table(
        [
            ["Usage charges total", _fmt_money(usage_total)],
            ["Access fee", _fmt_money(internal_fee)],
            ["Invoice total", _fmt_money(invoice_total)],
        ],
        colWidths=[3.0 * inch, 1.2 * inch],
        style=TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FCE4D6")),
            ]
        ),
        hAlign="LEFT",
    )
    story.append(totals_tbl)

    # Build PDF
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)


# -----------------------------
# Main pipeline
# -----------------------------
def load_and_prepare(
    csv_path: str,
    consumable_lab_map: Optional[Dict[str, str]] = None,
    *,
    tools_by_id: Optional[Dict[int, str]] = None,
    project_map: Optional[Dict[str, dict[str, Any]]] = None,
    adjustment_requests: Optional[list[dict[str, Any]]] = None,
) -> pd.DataFrame:
    df = pd.read_csv(csv_path)

    required = {
        "Type",
        "User",
        "Item",
        "Project",
        "Application identifier",
        "Start time",
        "Rate",
        "Cost",
        "Quantity",
    }
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV missing expected columns: {sorted(missing)}")

    # Include only supported application identifiers for invoicing
    df["Application identifier"] = df["Application identifier"].astype(str).str.strip()
    df = df[df["Application identifier"].isin(INVOICE_APPLICATION_IDENTIFIERS)].copy()

    df["Start_dt"] = df["Start time"].apply(parse_nemo_datetime)
    df["Item_norm"] = df["Item"].apply(normalize_item)
    df["IsConsumable"] = df["Type"].apply(_is_consumable_type)

    df["Lab"] = df["Item_norm"].map(TOOL_TO_LAB)
    if consumable_lab_map:
        df["Lab"] = df["Lab"].fillna(df["Item_norm"].map(consumable_lab_map))
    df["Lab"] = df["Lab"].fillna("Unmapped")
    df["Lab"] = df["Lab"].map(LAB_NAME_MAP).fillna(df["Lab"])
    df["Cost"] = pd.to_numeric(df["Cost"], errors="coerce").fillna(0.0).astype(float)
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").astype(float)

    if tools_by_id and adjustment_requests:
        df = apply_adjustment_requests(
            df,
            adjustment_requests=adjustment_requests,
            tools_by_id=tools_by_id,
            projects_by_name=project_map or {},
        )
        df["Item_norm"] = df["Item"].apply(normalize_item)
        df["IsConsumable"] = df["Type"].apply(_is_consumable_type)
        df["Lab"] = df["Item_norm"].map(TOOL_TO_LAB)
        if consumable_lab_map:
            df["Lab"] = df["Lab"].fillna(df["Item_norm"].map(consumable_lab_map))
        df["Lab"] = df["Lab"].fillna("Unmapped")
        df["Lab"] = df["Lab"].map(LAB_NAME_MAP).fillna(df["Lab"])

    df["Period"] = df["Start_dt"].apply(period_from_start_dt)
    df["Billable User Key"] = df.apply(resolve_billable_user_key, axis=1)
    df = apply_max_session_charge_caps(df)
    df = apply_project_charge_caps(df)
    df["Subsidy"] = 0.0
    cdg_mask = df["Application identifier"].str.upper().eq("CDG")
    if cdg_mask.any():
        # Show theoretical savings. The CSV cost is assumed already discounted (charged value).
        # If the discounted price = 90% of full price, savings is 1/9 of the charged cost.
        df.loc[cdg_mask, "Subsidy"] = df.loc[cdg_mask, "Cost"] / 9.0
        # If the charged amount is exactly the tool minimum, do not display a subsidy.
        minimum_charge = df.loc[cdg_mask, "Rate"].apply(parse_minimum_charge_from_rate)
        minimum_charge_mask = minimum_charge.notna() & (
            (df.loc[cdg_mask, "Cost"] - minimum_charge).abs() < 0.005
        )
        if minimum_charge_mask.any():
            df.loc[df.loc[cdg_mask].index[minimum_charge_mask], "Subsidy"] = 0.0
        # Do not adjust `Cost`; keep the value as in the CSV.

    return df


def generate_invoices(
    csv_path: str,
    outdir: str,
    nemo_base: Optional[str] = None,
    api_token: Optional[str] = None,
    generate_excel: bool = True,
    generate_pdf: bool = True,
    logo_path: Optional[str] = None,
    use_cache: bool = True,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    status_callback: Optional[Callable[[str], None]] = None,
) -> Tuple[int, int, pd.DataFrame, List[str]]:
    """
    Returns: (xlsx_created, pdf_created, prepared_df, generated_paths)
    """
    os.makedirs(outdir, exist_ok=True)

    if not generate_excel and not generate_pdf:
        raise RuntimeError("At least one output format must be selected.")

    if generate_pdf and not _pdf_available():
        raise RuntimeError(
            "PDF generation was requested, but reportlab is not installed."
        )

    project_map: Dict[str, dict] = {}
    consumable_lab_map: Dict[str, str] = {}
    tools_by_id: Dict[int, str] = {}
    adjustment_requests: list[dict[str, Any]] = []
    use_api = bool(nemo_base and api_token)

    if use_api:
        if status_callback:
            status_callback("Fetching consumable metadata from NEMO API")
        if progress_callback:
            progress_callback(0, 0, "Fetching NEMO consumable data")
        if requests is None:
            raise RuntimeError("requests is not installed; cannot use an API token.")
        consumable_lab_map = fetch_all_consumables(
            nemo_base=nemo_base,
            api_token=api_token,
            use_cache=use_cache,
            status_callback=status_callback,
        )
        if status_callback:
            status_callback("Fetching tools metadata from NEMO API")
        tools_by_id = fetch_all_tools(
            nemo_base=nemo_base,
            api_token=api_token,
            use_cache=use_cache,
            status_callback=status_callback,
        )
        if status_callback:
            status_callback("Fetching adjustment requests from NEMO API")
        adjustment_requests = fetch_all_adjustment_requests(
            nemo_base=nemo_base,
            api_token=api_token,
            use_cache=use_cache,
            status_callback=status_callback,
        )

    if status_callback:
        status_callback("Reading and preparing usage CSV")
    if progress_callback:
        progress_callback(0, 0, "Reading usage CSV")
    if use_api:
        if status_callback:
            status_callback("Fetching project contact data from NEMO API")
        if progress_callback:
            progress_callback(0, 0, "Fetching NEMO project contacts")
        project_map = fetch_all_projects(
            nemo_base=nemo_base,
            api_token=api_token,
            use_cache=use_cache,
            status_callback=status_callback,
        )

    df = load_and_prepare(
        csv_path,
        consumable_lab_map=consumable_lab_map,
        tools_by_id=tools_by_id,
        project_map=project_map,
        adjustment_requests=adjustment_requests,
    )
    if df.empty:
        return 0, 0, df

    if use_api:
        pi_infos = df["Project"].apply(
            lambda p: resolve_pi_for_project(str(p), project_map)
        )
        df["PI_key"] = pi_infos.apply(lambda x: x.key)
        df["PI_display_name"] = pi_infos.apply(lambda x: x.display_name)
        df["PI_email"] = pi_infos.apply(lambda x: x.email)
    else:
        df["PI_key"] = df["Project"].apply(extract_pi_code_from_project)
        df["PI_display_name"] = df["PI_key"]
        df["PI_email"] = ""

    xlsx_created = 0
    pdf_created = 0
    generated_paths: List[str] = []

    month_sequence: Dict[str, int] = {}
    grouped = df.groupby(["PI_key", "Period"], sort=True)
    total_invoices = grouped.ngroups
    processed_invoices = 0

    if status_callback:
        status_callback("Building PI contact summary workbook")
    contact_report_path = create_pi_contact_report(outdir, df)
    generated_paths.append(contact_report_path)

    if status_callback:
        status_callback(f"Prepared {total_invoices} invoice group(s)")
    if progress_callback:
        progress_callback(0, total_invoices, "Prepared invoice groups")

    for (pi_key, period), grp in grouped:
        pi_name = grp["PI_display_name"].iloc[0] or str(pi_key)
        nonempty_emails = grp["PI_email"].dropna().astype(str).str.strip()
        nonempty_emails = nonempty_emails[nonempty_emails != ""]
        pi_email = nonempty_emails.iloc[0] if not nonempty_emails.empty else ""
        period_key = str(period)
        month_sequence[period_key] = month_sequence.get(period_key, 0) + 1
        invoice_number = make_invoice_number(period_key, seq=month_sequence[period_key])

        filename_safe = safe_filename(pi_name)
        period_label = month_label(period)

        if generate_excel:
            if status_callback:
                status_callback(f"Building Excel for {pi_name} {period_label}")
            wb = create_invoice_workbook(
                grp,
                pi_display_name=pi_name,
                period_ym=period,
                invoice_number=invoice_number,
                pi_email=pi_email,
            )
            xlsx_path = os.path.join(outdir, f"{filename_safe} {period_label}.xlsx")
            wb.save(xlsx_path)
            generated_paths.append(xlsx_path)
            xlsx_created += 1
            try:
                wb.close()
            except Exception:
                pass

        # PDF (optional)
        if generate_pdf:
            if not _pdf_available():
                print(
                    "WARNING: reportlab is not installed; skipping PDF generation.",
                    file=sys.stderr,
                )
            else:
                pdf_path = os.path.join(outdir, f"{filename_safe} {period_label}.pdf")
                try:
                    if status_callback:
                        status_callback(f"Building PDF for {pi_name} {period_label}")
                    create_invoice_pdf(
                        grp,
                        pi_display_name=pi_name,
                        pi_email=pi_email,
                        period_ym=period,
                        invoice_number=invoice_number,
                        pdf_path=pdf_path,
                        logo_path=logo_path,
                    )
                    generated_paths.append(pdf_path)
                    pdf_created += 1
                except Exception as e:
                    if status_callback:
                        status_callback(
                            f"WARNING: PDF failed for {pi_name} {period_label}: {e}"
                        )
                    print(
                        f"WARNING: Failed to create PDF for {pi_name} {period_label}: {e}",
                        file=sys.stderr,
                    )
                    traceback.print_exc(file=sys.stderr)

        processed_invoices += 1
        if progress_callback:
            progress_callback(
                processed_invoices,
                total_invoices,
                f"{pi_name} {period_label}",
            )

    if generate_pdf and xlsx_created != pdf_created:
        print(
            f"WARNING: PDF generation incomplete: created {pdf_created} PDF(s) for {xlsx_created} invoice(s). "
            "XLSX files were still created.",
            file=sys.stderr,
        )

    return xlsx_created, pdf_created, df, generated_paths


def create_pi_contact_report(outdir: str, df: pd.DataFrame) -> str:
    month_labels = sorted({month_label(str(p)) for p in df["Period"].dropna().unique()})
    if len(month_labels) == 1:
        filename = f"CNI-Nemo-Invoice-PI-Contacts-{month_labels[0]}.xlsx"
    else:
        filename = (
            f"CNI-Nemo-Invoice-PI-Contacts-"
            f"{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        )

    report_df = (
        df.loc[:, ["PI_display_name", "PI_email", "Period"]]
        .copy()
        .rename(
            columns={
                "PI_display_name": "PI Name",
                "PI_email": "PI Email",
                "Period": "Billing Period",
            }
        )
    )
    report_df["PI Name"] = report_df["PI Name"].fillna("").astype(str).str.strip()
    report_df["PI Email"] = report_df["PI Email"].fillna("").astype(str).str.strip()
    report_df["Billing Period"] = (
        report_df["Billing Period"].fillna("").astype(str).str.strip()
    )
    report_df = report_df.drop_duplicates().sort_values(
        by=["PI Name", "Billing Period", "PI Email"], kind="stable"
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PI Contacts"
    worksheet.append(["PI Name", "PI Email", "Billing Period"])

    for row in report_df.itertuples(index=False):
        worksheet.append(list(row))

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 14), 48
        )

    output_path = os.path.join(outdir, filename)
    workbook.save(output_path)
    try:
        workbook.close()
    except Exception:
        pass
    return output_path


def create_invoice_zip(
    outdir: str, df: pd.DataFrame, remove_members: bool = True
) -> Optional[str]:
    """
    Create one ZIP containing all generated invoice files (.xlsx/.pdf) for the
    billing month(s) present in df. Optionally remove source files after zipping.
    """
    if df.empty or "Period" not in df.columns:
        return None

    month_labels = sorted({month_label(str(p)) for p in df["Period"].dropna().unique()})
    if not month_labels:
        return None

    members: List[str] = []
    for name in os.listdir(outdir):
        p = os.path.join(outdir, name)
        if not os.path.isfile(p):
            continue
        if not (name.endswith(".xlsx") or name.endswith(".pdf")):
            continue
        members.append(p)

    if not members:
        return None

    if len(month_labels) == 1:
        zip_name = f"CNI-Nemo-Invoices-{month_labels[0]}.zip"
    else:
        zip_name = (
            f"CNI-Nemo-Invoices-{invoice_generated_at().strftime('%Y%m%d-%H%M%S')}.zip"
        )
    zip_path = os.path.abspath(os.path.join(outdir, zip_name))

    with zipfile.ZipFile(zip_path, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for m in sorted(members):
            zf.write(m, arcname=os.path.basename(m))

    if remove_members:
        for m in members:
            try:
                os.remove(m)
            except OSError:
                pass

    return zip_path


# -----------------------------
# GUI helpers (tkinter)
# -----------------------------
def _pick_file_gui(title: str, filetypes):
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    root.update()
    path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return path or None


def _pick_csv_gui() -> Optional[str]:
    return _pick_file_gui(
        title="Select NEMO usage export CSV",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
    )


def _pick_logo_gui() -> Optional[str]:
    return _pick_file_gui(
        title="Select logo image for PDF (PNG/JPG). Cancel to skip.",
        filetypes=[("Image files", "*.png *.jpg *.jpeg"), ("All files", "*.*")],
    )


def _show_info_gui(title: str, message: str) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox
    except Exception:
        return
    root = tk.Tk()
    root.withdraw()
    root.update()
    messagebox.showinfo(title, message)
    root.destroy()


def _prompt_token_gui() -> Optional[str]:
    """Ask for a NEMO API token via GUI; returns token or None."""
    try:
        import tkinter as tk
        from tkinter import simpledialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    root.update()
    token = simpledialog.askstring(
        "NEMO API Token",
        "Enter your NEMO API token (leave blank to skip):",
        show="*",
    )
    root.destroy()
    if token is None:
        return None
    token = token.strip()
    return token or None


def _prompt_token_console() -> Optional[str]:
    """Ask for a NEMO API token in the terminal; returns token or None."""
    if not sys.stdin or not sys.stdin.isatty():
        return None
    try:
        token = getpass.getpass(
            "Enter your NEMO API token (leave blank to skip): "
        ).strip()
    except Exception:
        token = input("Enter your NEMO API token (leave blank to skip): ").strip()
    return token or None


def _get_api_token(no_gui: bool) -> Optional[str]:
    if no_gui:
        return _prompt_token_console()
    token = _prompt_token_gui()
    if token is not None:
        return token
    return _prompt_token_console()


def _default_logo_near_script() -> Optional[str]:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for fname in (
        "columbia_logo.png",
        "columbia_logo.jpg",
        "columbia_logo.jpeg",
        "logo.png",
        "logo.jpg",
        "logo.jpeg",
    ):
        cand = os.path.join(script_dir, fname)
        if os.path.exists(cand):
            return cand
    return None


def launch_gui_app() -> None:
    """Run a full GUI window for invoice generation."""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except Exception as e:
        raise RuntimeError(f"tkinter is required for GUI mode: {e}")

    root = tk.Tk()
    root.title("CNI NEMO Invoice Generator")
    root.geometry("880x430")
    root.minsize(820, 390)
    panel_bg = "#1F2937"
    text_fg = "#F9FAFB"
    entry_bg = "#F3F6FB"
    entry_fg = "#111827"
    root.configure(bg=panel_bg)

    csv_var = tk.StringVar(value="")
    outdir_var = tk.StringVar(value="")
    base_var = tk.StringVar(value=NEMO_BASE_URL)
    token_var = tk.StringVar(value="")
    logo_var = tk.StringVar(value="")
    pdf_var = tk.BooleanVar(value=_pdf_available())
    status_var = tk.StringVar(value="Ready.")

    def pick_csv():
        p = filedialog.askopenfilename(
            title="Select NEMO usage export CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if p:
            csv_var.set(p)
            if not outdir_var.get().strip():
                outdir_var.set(os.path.join(os.path.dirname(p), "invoices"))

    def pick_outdir():
        p = filedialog.askdirectory(title="Select output folder")
        if p:
            outdir_var.set(p)

    def pick_logo():
        p = filedialog.askopenfilename(
            title="Select logo image for PDF",
            filetypes=[("Image files", "*.png *.jpg *.jpeg"), ("All files", "*.*")],
        )
        if p:
            logo_var.set(p)

    def set_busy(is_busy: bool):
        state = "disabled" if is_busy else "normal"
        run_btn.config(state=state)
        for w in (csv_entry, outdir_entry, base_entry, token_entry, logo_entry):
            w.config(state=state)

    def run_job():
        csv_path = csv_var.get().strip()
        if not csv_path:
            messagebox.showerror("Missing CSV", "Please select a CSV file.")
            return
        if not os.path.exists(csv_path):
            messagebox.showerror("CSV not found", f"CSV not found:\n{csv_path}")
            return

        outdir = outdir_var.get().strip() or os.path.join(
            os.path.dirname(csv_path), "invoices"
        )
        outdir = os.path.abspath(outdir)
        os.makedirs(outdir, exist_ok=True)

        token = token_var.get().strip() or None
        nemo_base = (base_var.get().strip() or NEMO_BASE_URL).rstrip("/")

        generate_pdf_requested = bool(pdf_var.get())
        generate_pdf = generate_pdf_requested and _pdf_available()

        logo_path = logo_var.get().strip() or None
        if generate_pdf:
            if logo_path:
                logo_path = os.path.abspath(logo_path)
            else:
                logo_path = _default_logo_near_script()
        else:
            logo_path = None

        set_busy(True)
        status_var.set("Generating invoices...")
        root.update_idletasks()
        try:
            xlsx_created, pdf_created, df, _generated_paths = generate_invoices(
                csv_path=os.path.abspath(csv_path),
                outdir=outdir,
                nemo_base=nemo_base,
                api_token=token,
                generate_excel=True,
                generate_pdf=generate_pdf,
                logo_path=logo_path,
            )
            if generate_pdf and pdf_created == 0:
                raise RuntimeError(
                    "PDF generation was selected, but no PDFs were created. Check that reportlab is installed and that PDF generation did not hit an error."
                )
            zip_path = create_invoice_zip(outdir, df, remove_members=True)

            msg = f"Created {xlsx_created} XLSX invoice(s)"
            if generate_pdf_requested and not generate_pdf:
                msg += "\nPDF generation was selected, but reportlab is not installed. Only XLSX files were created."
            elif generate_pdf:
                msg += f" and {pdf_created} PDF invoice(s)"
            if zip_path:
                msg += f"\nCreated ZIP:\n{zip_path}"
                msg += "\n(Individual XLSX/PDF files were removed after ZIP creation.)"
            msg += f"\n\nOutput folder:\n{outdir}"
            status_var.set("Completed.")
            messagebox.showinfo("Done", msg)
        except Exception as e:
            status_var.set("Failed.")
            messagebox.showerror("Error", str(e))
        finally:
            set_busy(False)

    pad_x = 10
    pad_y = 7
    root.columnconfigure(1, weight=1)
    label_opts = {"bg": panel_bg, "fg": text_fg}
    check_opts = {
        "bg": panel_bg,
        "fg": text_fg,
        "selectcolor": panel_bg,
        "activebackground": panel_bg,
        "activeforeground": text_fg,
    }
    entry_opts = {
        "bg": entry_bg,
        "fg": entry_fg,
        "insertbackground": entry_fg,
        "disabledbackground": "#D1D5DB",
        "disabledforeground": "#4B5563",
    }

    tk.Label(root, text="NEMO CSV", **label_opts).grid(
        row=0, column=0, sticky="w", padx=pad_x, pady=pad_y
    )
    csv_entry = tk.Entry(root, textvariable=csv_var, **entry_opts)
    csv_entry.grid(row=0, column=1, sticky="ew", padx=pad_x, pady=pad_y)
    tk.Button(root, text="Browse...", command=pick_csv).grid(
        row=0, column=2, padx=pad_x, pady=pad_y
    )

    tk.Label(root, text="Output Folder", **label_opts).grid(
        row=1, column=0, sticky="w", padx=pad_x, pady=pad_y
    )
    outdir_entry = tk.Entry(root, textvariable=outdir_var, **entry_opts)
    outdir_entry.grid(row=1, column=1, sticky="ew", padx=pad_x, pady=pad_y)
    tk.Button(root, text="Browse...", command=pick_outdir).grid(
        row=1, column=2, padx=pad_x, pady=pad_y
    )

    tk.Label(root, text="NEMO Base URL", **label_opts).grid(
        row=2, column=0, sticky="w", padx=pad_x, pady=pad_y
    )
    base_entry = tk.Entry(root, textvariable=base_var, **entry_opts)
    base_entry.grid(row=2, column=1, sticky="ew", padx=pad_x, pady=pad_y)

    tk.Label(root, text="API Token", **label_opts).grid(
        row=3, column=0, sticky="w", padx=pad_x, pady=pad_y
    )
    token_entry = tk.Entry(root, textvariable=token_var, show="*", **entry_opts)
    token_entry.grid(row=3, column=1, sticky="ew", padx=pad_x, pady=pad_y)

    tk.Label(root, text="Logo (optional)", **label_opts).grid(
        row=4, column=0, sticky="w", padx=pad_x, pady=pad_y
    )
    logo_entry = tk.Entry(root, textvariable=logo_var, **entry_opts)
    logo_entry.grid(row=4, column=1, sticky="ew", padx=pad_x, pady=pad_y)
    tk.Button(root, text="Browse...", command=pick_logo).grid(
        row=4, column=2, padx=pad_x, pady=pad_y
    )

    pdf_checkbox = tk.Checkbutton(
        root, text="Generate PDF", variable=pdf_var, **check_opts
    )
    pdf_checkbox.grid(row=5, column=0, sticky="w", padx=pad_x, pady=pad_y)
    if not _pdf_available():
        pdf_checkbox.config(state="disabled")
        tk.Label(
            root, text="PDF unavailable: reportlab not installed", **label_opts
        ).grid(row=6, column=0, sticky="w", padx=pad_x, pady=pad_y)

    run_btn = tk.Button(root, text="Generate Invoices", command=run_job, width=22)
    run_btn.grid(row=7, column=0, padx=pad_x, pady=14, sticky="w")
    tk.Label(root, textvariable=status_var, anchor="w", **label_opts).grid(
        row=7, column=1, sticky="w", padx=pad_x, pady=14
    )

    root.mainloop()


# -----------------------------
# CLI entrypoint
# -----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--csv", required=False, default=None, help="Path to NEMO usage export CSV"
    )
    ap.add_argument(
        "--outdir",
        required=False,
        default=None,
        help='Directory to write invoice files into (default: "invoices" next to the CSV)',
    )
    ap.add_argument("--nemo-base", default=None, help=f"Base URL, e.g. {NEMO_BASE_URL}")
    ap.add_argument(
        "--api-token",
        default=None,
        help="NEMO API token (optional). If omitted, you will be prompted (leave blank to skip).",
    )
    ap.add_argument(
        "--logo",
        default=None,
        help="Path to a logo image (PNG/JPG) to place on the top-right of PDF invoices.",
    )
    ap.add_argument(
        "--no-pdf",
        action="store_true",
        help="Disable PDF generation (XLSX only).",
    )
    ap.add_argument(
        "--no-gui",
        action="store_true",
        help="Disable GUI file picker / popups (useful on servers).",
    )
    ap.add_argument(
        "--gui-app",
        action="store_true",
        help="Launch full desktop GUI app.",
    )
    args = ap.parse_args()

    if args.gui_app:
        launch_gui_app()
        return

    # CSV path
    csv_path = args.csv
    if not csv_path:
        if args.no_gui:
            raise SystemExit("Missing --csv (GUI disabled via --no-gui).")
        csv_path = _pick_csv_gui()
        if not csv_path:
            raise SystemExit("No CSV selected.")
    csv_path = os.path.abspath(csv_path)

    # Output folder
    outdir = args.outdir or os.path.join(os.path.dirname(csv_path), "invoices")
    outdir = os.path.abspath(outdir)
    os.makedirs(outdir, exist_ok=True)

    nemo_base = (args.nemo_base or NEMO_BASE_URL).rstrip("/")

    # Token (prompt if needed)
    token = args.api_token
    if token is None:
        token = _get_api_token(no_gui=args.no_gui)

    # Logo (optional; prompt if PDF enabled and GUI mode)
    # Logo (optional; used only for PDF generation)
    logo_path = args.logo
    if not args.no_pdf:
        if logo_path:
            logo_path = os.path.abspath(logo_path)
        else:
            # Try a default logo file next to this script (useful in --no-gui mode).
            logo_path = _default_logo_near_script()

        if not logo_path and not args.no_gui:
            # Let the user pick a logo (cancel is allowed)
            picked = _pick_logo_gui()
            if picked:
                logo_path = os.path.abspath(picked)

    generate_pdf = not args.no_pdf
    if generate_pdf and not _pdf_available():
        print(
            "WARNING: reportlab is not installed; generating XLSX only.",
            file=sys.stderr,
        )
        generate_pdf = False

    # Generate
    xlsx_created, pdf_created, df, _generated_paths = generate_invoices(
        csv_path=csv_path,
        outdir=outdir,
        nemo_base=nemo_base,
        api_token=(token or None),
        generate_excel=True,
        generate_pdf=generate_pdf,
        logo_path=logo_path,
    )
    zip_path = create_invoice_zip(outdir, df, remove_members=True)

    msg = f"Created {xlsx_created} XLSX invoice(s)"
    if generate_pdf:
        msg += f" and {pdf_created} PDF invoice(s)"
    if zip_path:
        msg += f"\nCreated ZIP:\n{zip_path}"
        msg += "\n(Individual XLSX/PDF files were removed after ZIP creation.)"
    msg += f" in:\n{outdir}"
    print(msg)

    if not args.no_gui and not args.csv:
        _show_info_gui("Invoices created", msg)


if __name__ == "__main__":
    main()
