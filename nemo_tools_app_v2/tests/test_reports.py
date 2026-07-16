from __future__ import annotations

import datetime as dt
import tempfile
import unittest
from pathlib import Path

import openpyxl

from nemo_app.reports.active_lab_users import build_active_lab_users_report
from nemo_app.reports.missed_reservations import build_missed_reservation_reports
from nemo_app.reports.usage_caps import build_usage_cap_report
from tests.fixtures import usage_frame


class _ActiveUsersClient:
    def fetch_all(self, endpoint: str):
        if endpoint == "tools/":
            return [{"id": 10, "name": "Clean Room SEM"}]
        if endpoint.startswith("qualifications/"):
            return [{"user": 20, "tool": 10, "qualified_on": "2026-06-01"}]
        raise AssertionError(f"Unexpected endpoint: {endpoint}")

    def fetch_by_ids(self, endpoint: str, identifiers):
        self.assert_user_lookup = (endpoint, set(identifiers))
        return {
            20: {
                "id": 20,
                "username": "al1",
                "email": "ada@example.edu",
                "first_name": "Ada",
                "last_name": "Lovelace",
                "is_active": True,
            }
        }


class ReportTests(unittest.TestCase):
    def test_usage_cap_workbook_keeps_source_labels_and_staff_audit(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "April source.csv"
            output = root / "cap-report.xlsx"
            usage_frame().to_csv(source, index=False)

            result = build_usage_cap_report([(source, "April Billing")], output)

            self.assertEqual(result.periods, ("2026-04",))
            self.assertGreater(result.final_income, 0)
            workbook = openpyxl.load_workbook(output, data_only=True)
            try:
                self.assertIn("Staff Charge Detail", workbook.sheetnames)
                assumptions = workbook["Assumptions"]
                values = [cell.value for row in assumptions.iter_rows() for cell in row]
                self.assertIn("April Billing", values)
                self.assertIn(source.name, values)
                self.assertNotIn(str(source), values)
            finally:
                workbook.close()

    def test_missed_reservations_and_active_users(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "usage.csv"
            usage_frame().to_csv(source, index=False)
            users, tools, total = build_missed_reservation_reports(source, threshold=1)
            self.assertEqual(total, 1)
            self.assertEqual(users.iloc[0]["User"], "Grace Hopper")
            self.assertEqual(tools.iloc[0]["Missed Reservations"], 1)

            output = root / "active.xlsx"
            active = build_active_lab_users_report(
                output,
                client=_ActiveUsersClient(),  # type: ignore[arg-type]
                selected_labs=("clean-room",),
                today=dt.date(2026, 7, 1),
            )
            self.assertEqual(active.combined_user_count, 1)
            self.assertTrue(output.is_file())


if __name__ == "__main__":
    unittest.main()
