from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from nemo_app.billing.invoice_model import InvoiceDocument
from nemo_app.billing.prepare import prepare_usage_dataframe
from nemo_app.invoices.excel_renderer import render_invoice_workbook
from nemo_app.invoices.pdf_renderer import render_invoice_pdf
from tests.fixtures import usage_frame


class InvoiceCharacterizationTests(unittest.TestCase):
    def document(self) -> InvoiceDocument:
        prepared = prepare_usage_dataframe(usage_frame(), apply_caps=False)
        prepared = prepared.loc[~prepared["IsMissedReservation"]].iloc[:3].copy()
        return InvoiceDocument.from_frame(
            prepared,
            pi_key="ada.pi@example.edu",
            pi_name="PI, Ada",
            pi_email="ada.pi@example.edu",
            period="2026-04",
            invoice_number="CNI-2604-TEST-001",
        )

    def test_document_calculates_totals_once(self) -> None:
        document = self.document()
        self.assertEqual(document.usage_total, 175.0)
        self.assertEqual(document.access_fee, 50.0)
        self.assertEqual(document.invoice_total, 225.0)
        self.assertEqual(sum(row.total for row in document.projects), 225.0)

    def test_excel_and_pdf_render_from_same_document(self) -> None:
        document = self.document()
        with tempfile.TemporaryDirectory() as folder:
            folder_path = Path(folder)
            xlsx_path = folder_path / "invoice.xlsx"
            pdf_path = folder_path / "invoice.pdf"
            render_invoice_workbook(document, xlsx_path)
            render_invoice_pdf(document, pdf_path)

            workbook = load_workbook(xlsx_path, data_only=False)
            try:
                sheet = workbook["Invoice"]
                self.assertEqual(sheet["B4"].value, document.pi_name)
                self.assertEqual(sheet["K4"].value, document.invoice_number)
                self.assertIn(
                    "Invoice total",
                    [sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)],
                )
            finally:
                workbook.close()
            self.assertGreater(pdf_path.stat().st_size, 1000)


if __name__ == "__main__":
    unittest.main()
