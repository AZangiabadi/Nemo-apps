from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from nemo_app.billing.constants import DESIRED_LAB_ORDER, DETAIL_SECTION_ORDER
from nemo_app.billing.invoice_model import InvoiceDocument
from nemo_app.billing.text import month_label

from .excel_styles import (
    BOLD,
    BORDER,
    CURRENCY_FORMAT,
    SECTION_FILL,
    TITLE,
    TITLE_LARGE,
    autosize_columns,
    style_table_header,
)


def _write_detail_table(worksheet, row: int, frame: pd.DataFrame, show_subsidy: bool) -> int:
    columns = ["Date", "User", "Description", "Type", "Project", "Quantity", "Rate"]
    if show_subsidy:
        columns.append("Subsidy")
    columns.append("Cost")
    style_table_header(worksheet, row, 1, columns)
    for _, source in frame.iterrows():
        row += 1
        values = [
            source.get("Start_dt"),
            source.get("User", ""),
            source.get("Item_norm", ""),
            source.get("Type", ""),
            source.get("Project", ""),
            source.get("Quantity"),
            source.get("Rate", ""),
        ]
        if show_subsidy:
            values.append(float(source.get("Subsidy", 0.0) or 0.0))
        values.append(float(source.get("Cost", 0.0) or 0.0))
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if columns[column - 1] in {"Subsidy", "Cost"}:
                cell.number_format = CURRENCY_FORMAT
    row += 1
    worksheet.cell(row, len(columns) - 1, "Subtotal").font = BOLD
    subtotal = worksheet.cell(row, len(columns), float(frame["Cost"].sum()))
    subtotal.font = BOLD
    subtotal.number_format = CURRENCY_FORMAT
    return row + 2


def render_invoice_workbook(document: InvoiceDocument, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoice"
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells("A1:K1")
    worksheet["A1"] = "Columbia Nano Initiative"
    worksheet["A1"].font = TITLE_LARGE
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet.merge_cells("A2:K2")
    worksheet["A2"] = "Facility Usage Invoice"
    worksheet["A2"].font = TITLE
    worksheet["A2"].alignment = Alignment(horizontal="center")
    header = {
        "A4": "PI",
        "B4": document.pi_name,
        "D4": "Billing Month",
        "E4": month_label(document.period),
        "G4": "Generated",
        "H4": document.generated_at.replace(tzinfo=None),
        "J4": "Invoice #",
        "K4": document.invoice_number,
        "A5": "Email",
        "B5": document.pi_email or "N/A",
    }
    for coordinate, value in header.items():
        worksheet[coordinate] = value
    for coordinate in ("A4", "D4", "G4", "J4", "A5"):
        worksheet[coordinate].font = BOLD
    worksheet["B4"].font = worksheet["E4"].font = worksheet["K4"].font = TITLE
    worksheet["H4"].number_format = "yyyy-mm-dd hh:mm"

    row = 6
    style_table_header(worksheet, row, 1, ["Lab", "Total Cost"])
    for lab, total in sorted(document.lab_totals.items()):
        row += 1
        worksheet.cell(row, 1, lab).border = BORDER
        cell = worksheet.cell(row, 2, total)
        cell.border = BORDER
        cell.number_format = CURRENCY_FORMAT
    row += 1
    worksheet.cell(row, 1, "Access fee").border = BORDER
    worksheet.cell(row, 2, document.access_fee).number_format = CURRENCY_FORMAT
    row += 1
    worksheet.cell(row, 1, "TOTAL").font = BOLD
    worksheet.cell(row, 2, document.invoice_total).number_format = CURRENCY_FORMAT
    worksheet.cell(row, 2).font = BOLD
    row += 2

    detail_end_column = 9 if document.show_subsidy else 8
    for lab in DETAIL_SECTION_ORDER:
        lines = document.lines_for_lab(lab)
        if lines.empty:
            continue
        worksheet.cell(row, 1, lab).font = Font(bold=True, size=12)
        worksheet.cell(row, 1).fill = SECTION_FILL
        worksheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=detail_end_column
        )
        row = _write_detail_table(worksheet, row + 1, lines, document.show_subsidy)

    if document.access_fee_project:
        worksheet.cell(row, 1, "Access fee").font = Font(bold=True, size=12)
        worksheet.cell(row, 1).fill = SECTION_FILL
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        style_table_header(worksheet, row, 1, ["Project", "Project Type", "Access Fee"])
        row += 1
        worksheet.cell(row, 1, document.access_fee_project[0]).border = BORDER
        worksheet.cell(row, 2, document.access_fee_project[1]).border = BORDER
        worksheet.cell(row, 3, document.access_fee).number_format = CURRENCY_FORMAT
        worksheet.cell(row, 3).border = BORDER
        row += 2

    worksheet.cell(row, 1, "Project fees summary").font = Font(bold=True, size=12)
    worksheet.cell(row, 1).fill = SECTION_FILL
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    project_headers = [
        "Project",
        "Project Type",
        *DESIRED_LAB_ORDER,
        "Staff Time",
        "Access Fee",
        "Project Total",
    ]
    style_table_header(worksheet, row, 1, list(project_headers))
    for project in sorted(document.projects, key=lambda value: (-value.total, value.project)):
        row += 1
        values = [
            project.project,
            project.application,
            *(project.lab_totals[lab] for lab in DESIRED_LAB_ORDER),
            project.staff_time,
            project.access_fee,
            project.total,
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row, column, value)
            cell.border = BORDER
            if column >= 3:
                cell.number_format = CURRENCY_FORMAT

    last_column = len(project_headers)
    for label, value, bold in (
        ("Usage charges total", document.usage_total, True),
        ("Access fee", document.access_fee, False),
        ("Invoice total", document.invoice_total, True),
    ):
        row += 1
        worksheet.cell(row, 1, label).font = BOLD if bold else Font()
        worksheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=last_column - 1
        )
        cell = worksheet.cell(row, last_column, value)
        cell.number_format = CURRENCY_FORMAT
        if bold:
            cell.font = BOLD

    worksheet.freeze_panes = "A7"
    autosize_columns(worksheet)
    workbook.save(output_path)
    workbook.close()
    return output_path
