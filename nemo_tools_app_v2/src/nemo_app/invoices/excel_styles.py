from __future__ import annotations

import datetime as dt

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
TITLE_LARGE = Font(bold=True, size=18)
CURRENCY_FORMAT = "$#,##0.00;[Red]($#,##0.00)"


def autosize_columns(worksheet, *, minimum: int = 10, maximum: int = 60) -> None:
    for column in range(1, worksheet.max_column + 1):
        lengths = []
        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row, column).value
            if isinstance(value, (dt.date, dt.datetime)):
                lengths.append(len(value.strftime("%Y-%m-%d %H:%M")))
            elif value is not None:
                lengths.append(len(str(value)))
        width = max(lengths, default=minimum - 2) + 2
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            maximum, max(minimum, width)
        )


def style_table_header(worksheet, row: int, start_column: int, names: list[str]) -> None:
    for offset, name in enumerate(names):
        cell = worksheet.cell(row, start_column + offset, name)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
