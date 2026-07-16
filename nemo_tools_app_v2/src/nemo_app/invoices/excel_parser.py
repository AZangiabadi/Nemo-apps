from __future__ import annotations

import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nemo_app.billing.constants import DESIRED_LAB_ORDER
from nemo_app.billing.invoice_model import InvoiceDocument
from nemo_app.billing.text import month_label, parse_nemo_datetime, safe_filename

from .pdf_renderer import render_invoice_pdf


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace("$", "").replace(",", ""))
    except ValueError:
        return 0.0


def _period(value: object) -> str:
    text = _text(value)
    for format_string in ("%b%Y", "%B%Y", "%b %Y", "%B %Y", "%Y-%m"):
        try:
            return dt.datetime.strptime(text, format_string).strftime("%Y-%m")
        except ValueError:
            continue
    return text or dt.date.today().strftime("%Y-%m")


def parse_invoice_workbook(path: Path) -> InvoiceDocument:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["Invoice"] if "Invoice" in workbook.sheetnames else workbook.active
        project_types: dict[str, str] = {}
        for row in range(1, sheet.max_row + 1):
            if _text(sheet.cell(row, 1).value) != "Project fees summary":
                continue
            headers = {
                _text(sheet.cell(row + 1, column).value): column
                for column in range(1, sheet.max_column + 1)
            }
            project_column = headers.get("Project")
            type_column = headers.get("Project Type")
            if project_column and type_column:
                for source_row in range(row + 2, sheet.max_row + 1):
                    project = _text(sheet.cell(source_row, project_column).value)
                    if not project:
                        break
                    project_types[project] = _text(sheet.cell(source_row, type_column).value)
            break

        records: list[dict[str, object]] = []
        row = 1
        sections = {*DESIRED_LAB_ORDER, "Staff time"}
        while row <= sheet.max_row:
            lab = _text(sheet.cell(row, 1).value)
            if lab not in sections:
                row += 1
                continue
            headers = {
                _text(sheet.cell(row + 1, column).value): column
                for column in range(1, sheet.max_column + 1)
            }
            required = {
                "Date",
                "User",
                "Description",
                "Type",
                "Project",
                "Quantity",
                "Rate",
                "Cost",
            }
            if not required.issubset(headers):
                row += 1
                continue
            source_row = row + 2
            while source_row <= sheet.max_row:
                first = _text(sheet.cell(source_row, 1).value)
                values = [
                    _text(sheet.cell(source_row, column).value)
                    for column in range(1, sheet.max_column + 1)
                ]
                if (
                    first in sections
                    or first in {"Access fee", "Project fees summary"}
                    or "Subtotal" in values
                ):
                    break
                project = _text(sheet.cell(source_row, headers["Project"]).value)
                item = _text(sheet.cell(source_row, headers["Description"]).value)
                user = _text(sheet.cell(source_row, headers["User"]).value)
                if project or item or user:
                    start = sheet.cell(source_row, headers["Date"]).value
                    start_dt = (
                        start if isinstance(start, dt.datetime) else parse_nemo_datetime(start)
                    )
                    type_value = _text(sheet.cell(source_row, headers["Type"]).value)
                    is_consumable = lab == "Consumable"
                    is_staff = item.lower() == "staff time" or type_value.lower() == "staff_charge"
                    records.append(
                        {
                            "Type": type_value,
                            "User": user,
                            "Username": "",
                            "Item": item,
                            "Item_norm": item,
                            "Project": project,
                            "Application identifier": project_types.get(project, "Local"),
                            "Start_dt": start_dt,
                            "End_dt": None,
                            "Quantity": _number(sheet.cell(source_row, headers["Quantity"]).value),
                            "Rate": _text(sheet.cell(source_row, headers["Rate"]).value),
                            "Cost": _number(sheet.cell(source_row, headers["Cost"]).value),
                            "Subsidy": _number(
                                sheet.cell(source_row, headers.get("Subsidy", 0)).value
                            )
                            if headers.get("Subsidy")
                            else 0.0,
                            "Lab": lab,
                            "IsConsumable": is_consumable,
                            "IsStaffCharge": is_staff,
                            "IsToolUsageCharge": not is_consumable and not is_staff,
                            "IsMissedReservation": False,
                        }
                    )
                source_row += 1
            row = source_row
        if not records:
            raise ValueError("No invoice detail lines were found in the workbook.")
        period = _period(sheet["E4"].value)
        frame = pd.DataFrame(records)
        frame["Period"] = period
        frame["Billable User Key"] = frame["User"]
        return InvoiceDocument.from_frame(
            frame,
            pi_key=_text(sheet["B5"].value),
            pi_name=_text(sheet["B4"].value) or "UNKNOWN_PI",
            pi_email="" if _text(sheet["B5"].value) == "N/A" else _text(sheet["B5"].value),
            period=period,
            invoice_number=_text(sheet["K4"].value) or f"CNI-{period.replace('-', '')}",
            access_fee_override=_read_access_fee(sheet),
        )
    finally:
        workbook.close()


def _read_access_fee(sheet) -> float:
    for row in range(1, sheet.max_row + 1):
        if _text(sheet.cell(row, 1).value).lower() != "access fee":
            continue
        if _text(sheet.cell(row + 1, 3).value) == "Access Fee":
            return _number(sheet.cell(row + 2, 3).value)
        return _number(sheet.cell(row, 2).value)
    return 0.0


def convert_excel_to_pdf(
    workbook_path: Path,
    output_dir: Path,
    *,
    logo_path: Path | None = None,
) -> Path:
    document = parse_invoice_workbook(workbook_path)
    output_path = output_dir / (
        f"{safe_filename(document.pi_name)} {month_label(document.period)}.pdf"
    )
    return render_invoice_pdf(document, output_path, logo_path=logo_path)
