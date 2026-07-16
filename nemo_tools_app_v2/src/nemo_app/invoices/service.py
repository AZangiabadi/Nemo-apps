from __future__ import annotations

import datetime as dt
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from nemo_app.billing.invoice_model import InvoiceDocument
from nemo_app.billing.prepare import load_usage_csv
from nemo_app.billing.text import month_label, safe_filename
from nemo_app.nemo.metadata import MetadataRepository, project_pi

from .excel_renderer import render_invoice_workbook
from .excel_styles import CURRENCY_FORMAT, autosize_columns, style_table_header
from .pdf_renderer import render_invoice_pdf

ProgressCallback = Callable[[int, int, str], None]


@dataclass(frozen=True, slots=True)
class InvoiceOptions:
    generate_excel: bool = True
    generate_pdf: bool = True
    make_zip: bool = True
    use_cache: bool = True
    apply_hourly_caps: bool = True


@dataclass(frozen=True, slots=True)
class InvoiceResult:
    files: tuple[Path, ...]
    invoice_count: int
    excel_count: int
    pdf_count: int
    row_count: int


def make_invoice_number(period: str, sequence: int, generated_at: dt.datetime) -> str:
    period_code = period.replace("-", "")[-4:]
    return f"CNI-{period_code}-{generated_at.strftime('%d%H%M')}-{sequence:03d}"


def _has_invoiceable_activity(frame: pd.DataFrame) -> bool:
    if frame.empty or float(frame["Cost"].fillna(0).sum()) <= 0:
        return False
    return bool(
        frame["IsToolUsageCharge"].fillna(False).any() or frame["IsConsumable"].fillna(False).any()
    )


def _contact_report(documents: list[InvoiceDocument], path: Path) -> Path:
    workbook = Workbook()
    totals_sheet = workbook.active
    totals_sheet.title = "PI Project Totals"
    headers = [
        "PI Name",
        "PI Email",
        "Billing Period",
        "Project Number",
        "Project Type",
        "Usage Amount",
        "Access Fee",
        "Total Amount",
    ]
    style_table_header(totals_sheet, 1, 1, headers)
    row = 1
    for document in documents:
        for project in document.projects:
            row += 1
            values = [
                document.pi_name,
                document.pi_email,
                document.period,
                project.project,
                project.application,
                project.usage_total,
                project.access_fee,
                project.total,
            ]
            for column, value in enumerate(values, 1):
                totals_sheet.cell(row, column, value)
                if column >= 6:
                    totals_sheet.cell(row, column).number_format = CURRENCY_FORMAT
    autosize_columns(totals_sheet, minimum=14, maximum=56)

    users_sheet = workbook.create_sheet("Tool Users")
    style_table_header(users_sheet, 1, 1, ["User", "Project Number", "Project Type"])
    seen: set[tuple[str, str, str]] = set()
    for document in documents:
        tool_lines = document.lines.loc[document.lines["IsToolUsageCharge"].fillna(False)]
        for _, line in tool_lines.iterrows():
            record = (
                str(line.get("User") or "").strip(),
                str(line.get("Project") or "").strip(),
                str(line.get("Application identifier") or "").strip(),
            )
            if record[0]:
                seen.add(record)
    for row_index, record in enumerate(sorted(seen), 2):
        for column, value in enumerate(record, 1):
            users_sheet.cell(row_index, column, value)
    autosize_columns(users_sheet, minimum=14, maximum=56)
    workbook.save(path)
    workbook.close()
    return path


def _archive(paths: list[Path], output_path: Path) -> Path:
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(paths):
            archive.write(path, arcname=path.name)
    return output_path


def generate_invoices(
    csv_path: Path,
    output_dir: Path,
    *,
    metadata: MetadataRepository,
    options: InvoiceOptions,
    logo_path: Path | None = None,
    progress: ProgressCallback | None = None,
) -> InvoiceResult:
    if not options.generate_excel and not options.generate_pdf:
        raise ValueError("Select at least one invoice output format.")
    output_dir.mkdir(parents=True, exist_ok=True)
    projects = metadata.projects(use_cache=options.use_cache)
    tools = metadata.tools(use_cache=options.use_cache)
    adjustments = metadata.adjustments(use_cache=options.use_cache)
    consumables = metadata.consumable_labs(use_cache=options.use_cache)
    prepared = load_usage_csv(
        csv_path,
        consumable_labs=consumables,
        tools_by_id=tools,
        project_map=projects,
        adjustment_requests=adjustments,
        filter_invoice_quantities=True,
        apply_hourly_caps=options.apply_hourly_caps,
    )
    prepared = prepared.loc[~prepared["IsMissedReservation"].fillna(False)].copy()
    if prepared.empty:
        raise ValueError("No invoiceable rows remain after filtering the CSV.")
    pi_values = prepared["Project"].apply(lambda value: project_pi(str(value), projects))
    prepared["PI_key"] = pi_values.apply(lambda value: value.key)
    prepared["PI_name"] = pi_values.apply(lambda value: value.display_name)
    prepared["PI_email"] = pi_values.apply(lambda value: value.email)

    groups = [
        (key, group.copy())
        for key, group in prepared.groupby(["PI_key", "Period"], sort=True)
        if _has_invoiceable_activity(group)
    ]
    if not groups:
        raise ValueError("No invoice groups contain billable activity.")
    generated_at = dt.datetime.now().astimezone()
    sequences: dict[str, int] = {}
    documents: list[InvoiceDocument] = []
    used_names: set[str] = set()
    files: list[Path] = []
    excel_count = pdf_count = 0
    for index, ((pi_key, period), group) in enumerate(groups, 1):
        period = str(period)
        sequences[period] = sequences.get(period, 0) + 1
        name = str(group["PI_name"].iloc[0] or pi_key)
        email_values = group["PI_email"].fillna("").astype(str).str.strip()
        email = email_values.loc[email_values.ne("")].iloc[0] if email_values.ne("").any() else ""
        document = InvoiceDocument.from_frame(
            group,
            pi_key=str(pi_key),
            pi_name=name,
            pi_email=email,
            period=period,
            invoice_number=make_invoice_number(period, sequences[period], generated_at),
            generated_at=generated_at,
        )
        documents.append(document)
        base_name = f"{safe_filename(name)} {month_label(period)}"
        candidate = base_name
        suffix = 0
        while candidate in used_names:
            suffix += 1
            candidate = f"{base_name}_{suffix}"
        used_names.add(candidate)
        if options.generate_excel:
            files.append(render_invoice_workbook(document, output_dir / f"{candidate}.xlsx"))
            excel_count += 1
        if options.generate_pdf:
            files.append(
                render_invoice_pdf(document, output_dir / f"{candidate}.pdf", logo_path=logo_path)
            )
            pdf_count += 1
        if progress:
            progress(index, len(groups), f"{name} {month_label(period)}")

    periods = sorted({document.period for document in documents})
    contact_name = (
        f"CNI-Nemo-Invoice-PI-Contacts-{month_label(periods[0])}.xlsx"
        if len(periods) == 1
        else f"CNI-Nemo-Invoice-PI-Contacts-{generated_at.strftime('%Y%m%d-%H%M%S')}.xlsx"
    )
    files.append(_contact_report(documents, output_dir / contact_name))
    if options.make_zip:
        archive_name = (
            f"CNI-Nemo-Invoices-{month_label(periods[0])}.zip"
            if len(periods) == 1
            else f"CNI-Nemo-Invoices-{generated_at.strftime('%Y%m%d-%H%M%S')}.zip"
        )
        archive = _archive(files, output_dir / archive_name)
        for path in files:
            path.unlink(missing_ok=True)
        files = [archive]
    return InvoiceResult(
        files=tuple(files),
        invoice_count=len(documents),
        excel_count=excel_count,
        pdf_count=pdf_count,
        row_count=len(prepared),
    )
